#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SkyProfile (PyQt5, consolidated ADSBx flight history exports with enriched baseball card)

Features
--------
- Query by date range + ICAO hex OR by tail/registration (via ADSBx basic-ac-db).
- Output KML (points, 2D routes, 3D routes – leg-based, not full traces),
  CSV, JSON (consolidated across days), and Excel summary.
- Output folder picker with "Open" to launch Explorer/Finder/xdg-open.
- Baseball card (right pane):
    * Aircraft type (ICAO code, e.g. GLF4)
    * Type name (friendly, e.g. Gulfstream IV / G-IV)
    * Registration (tail)
    * Registered owner
    * Manufacturer / model
    * Flags (Military / LADD / PIA / Interesting)
    * Callsigns seen
    * ICAO Hex
    * Aircraft image (Planespotters if available)

Data details
------------
- Downloads ADSBexchange globe_history trace_full JSONs.
- Parses segments & points (with timestamps, alt, gs, track, etc.).
- Builds:
    * KML:
        - (leg-based) Points, 2D routes, and 3D routes with per-leg metadata.
    * CSV:
        - Base columns (hex, time, lat, lon, alt, etc.)
        - One column per AC data key (flattened).
    * JSON:
        - Raw ADSBx trace_full blobs merged (no additional parsing).
    * Excel summary:
        - Multi-sheet workbook including Summary, Flights, Airports, Routes,
          DayOfWeek, DayOfMonth, Hours, Countries, Callsigns.

Attribution & Copyright
-----------------------
- ADSBexchange globe history data © ADSBexchange; used under their terms for personal/analytical use.
- ADSBexchange basic-ac-db © ADSBexchange; redistribution terms apply—see their documentation.
- OpenSky aircraft metadata © OpenSky Network contributors; API terms apply.
- Planespotters photo API © Planespotters.net and respective photographers; subject to their licensing/terms.
- OurAirports data © contributors to OurAirports (CC0/Public Domain as published); see https://ourairports.com/.
- Icons/images (if bundled) retain their original licenses; ensure compliance if redistributing.

Requirements
------------
    pip install PyQt5 requests simplekml openpyxl
"""

import os
import shutil
import sys
import json
import gzip
import io
import math
import time
import csv
import zipfile
import subprocess
import re
import unicodedata
import datetime as dt
import html
from dataclasses import dataclass, field
from typing import List, Dict, Any, Iterable, Optional, Set, Tuple

import requests
import collections
import simplekml
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.axis import ChartLines
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as XLImage

try:
    import pycountry  # optional, for nicer country names
except ImportError:
    pycountry = None

# Fallback country names for common ISO digraphs (used if pycountry is unavailable)
ISO_COUNTRY_FALLBACK = {
    "US": "United States",
    "CA": "Canada",
    "MX": "Mexico",
    "GB": "United Kingdom",
    "UK": "United Kingdom",
    "FR": "France",
    "DE": "Germany",
    "ES": "Spain",
    "IT": "Italy",
    "NL": "Netherlands",
    "BE": "Belgium",
    "CH": "Switzerland",
    "AT": "Austria",
    "AU": "Australia",
    "NZ": "New Zealand",
    "BR": "Brazil",
    "AR": "Argentina",
    "CL": "Chile",
    "ZA": "South Africa",
    "RU": "Russia",
    "CN": "China",
    "JP": "Japan",
    "KR": "South Korea",
    "SG": "Singapore",
    "IN": "India",
    "AE": "United Arab Emirates",
    "SA": "Saudi Arabia",
    "QA": "Qatar",
    "KW": "Kuwait",
    "OM": "Oman",
    "TR": "Turkey",
    "GR": "Greece",
    "SE": "Sweden",
    "NO": "Norway",
    "FI": "Finland",
    "DK": "Denmark",
    "IS": "Iceland",
    "IE": "Ireland",
    "PT": "Portugal",
    "PL": "Poland",
    "CZ": "Czech Republic",
    "HU": "Hungary",
    "RO": "Romania",
    "BG": "Bulgaria",
    "HR": "Croatia",
    "SI": "Slovenia",
    "SK": "Slovakia",
    "LT": "Lithuania",
    "LV": "Latvia",
    "EE": "Estonia",
    "IL": "Israel",
    "GY": "Guyana",
    "CO": "Colombia",
    "PE": "Peru",
    "CI": "Côte d'Ivoire",
    "GH": "Ghana",
    "GW": "Guinea-Bissau",
    "CV": "Cabo Verde",
    "JM": "Jamaica",
    "PA": "Panama",
    "VG": "Virgin Islands (British)",
    "EC": "Ecuador",
    "GL": "Greenland",
    "AI": "Anguilla",
    "BM": "Bermuda",
    "BO": "Bolivia",
    "FO": "Faroe Islands",
    "SR": "Suriname",
}


def iso_country_name(code: str) -> str:
    """Return human-friendly country name from ISO digraph. If the value already looks like a name, return it."""
    if not code:
        return ""
    raw = (code or "").strip()
    # If it's already a name (not a 2-letter code), keep it as-is
    if len(raw) > 2 and not raw.isupper():
        return raw

    code_upper = raw.upper()
    if pycountry:
        try:
            c = pycountry.countries.get(alpha_2=code_upper)
            if c and getattr(c, "name", None):
                return c.name
        except Exception:
            pass
    # Fallback
    if code_upper in ISO_COUNTRY_FALLBACK:
        return ISO_COUNTRY_FALLBACK[code_upper]
    return raw

from PyQt5 import QtCore, QtGui, QtWidgets


# -----------------------------
# Resource helpers (PyInstaller-safe)
# -----------------------------


def resource_path(rel_path: str) -> str:
    """
    Resolve a relative resource path that works both from source and PyInstaller onefile.
    """
    base_path = getattr(sys, "_MEIPASS", os.path.abspath("."))
    candidate = os.path.join(base_path, rel_path)
    if os.path.exists(candidate):
        return candidate
    # Fallback to resources/ subfolder
    alt = os.path.join(base_path, "resources", os.path.basename(rel_path))
    return alt


# -----------------------------
# Excel-safe text helper
# -----------------------------


_INVALID_CTRL_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def clean_cell_value(val: Any) -> Any:
    """Strip Excel-invalid control characters from strings."""
    if val is None:
        return ""
    if isinstance(val, str):
        return _INVALID_CTRL_RE.sub("", val)
    return val


# -----------------------------
# Misc helpers
# -----------------------------


def normalize_callsign(value: Any) -> str:
    """
    Normalize callsigns so weird spacing / Unicode / zero-width chars don't
    cause grouping bugs, but keep the underlying text as intact as possible.
    """
    if not isinstance(value, str):
        return ""

    s = unicodedata.normalize("NFKC", value)

    # Normalize all whitespace to plain spaces
    s = "".join(" " if ch.isspace() else ch for ch in s)

    # Strip control / zero-width characters
    s = "".join(
        ch
        for ch in s
        if not (unicodedata.category(ch).startswith("C") and ch not in ("\n", "\r", "\t"))
    )

    # Collapse repeated spaces and trim
    s = re.sub(r"\s+", " ", s).strip()

    return s


def daterange(start_date: dt.date, end_date: dt.date) -> Iterable[dt.date]:
    """Yield dates from start_date to end_date inclusive."""
    step = 1 if end_date >= start_date else -1
    cur = start_date
    while True:
        yield cur
        if cur == end_date:
            break
        cur = cur + dt.timedelta(days=step)


def ensure_dir_for_file(path: str) -> None:
    """Ensure parent directory exists for a file path."""
    d = os.path.dirname(os.path.abspath(path))
    if d and not os.path.exists(d):
        os.makedirs(d, exist_ok=True)


# -----------------------------
# ADSBx constants & helpers
# -----------------------------

HEADERS = {
    "Referer": "https://globe.adsbexchange.com/",
    "User-Agent": "adsbx-history-downloader-gui/pyqt/2.0",
}

BASE = (
    "https://globe.adsbexchange.com/globe_history/{y}/{m:02d}/{d:02d}/"
    "traces/{suffix}/trace_full_{hex}.json"
)

AC_DB_URL = "http://downloads.adsbexchange.com/downloads/basic-ac-db.json.gz"
_ACDB_CACHE = None  # in-memory cache of ADSBx aircraft DB

# ICAO aircraft type designator → common name
ICAO_TYPE_NAMES: Dict[str, str] = {
    "GLF4": "Gulfstream IV / G-IV",
    "GLF5": "Gulfstream V / G-V",
    "GLF6": "Gulfstream G650 / G6",
    # Extend as needed
}


# -----------------------------
# Dataclasses
# -----------------------------


@dataclass
class AircraftMeta:
    hex: str
    registration: Optional[str] = None
    type: Optional[str] = None
    type_name: Optional[str] = None
    owner: Optional[str] = None
    manufacturer: Optional[str] = None
    model: Optional[str] = None
    country: Optional[str] = None
    flags: Optional[str] = None
    callsigns: Optional[List[str]] = field(default_factory=list)
    description: Optional[str] = None
    photo_url: Optional[str] = None
    faa_data: Optional[Dict[str, Any]] = None


@dataclass
class OpenSkyMeta:
    registration: Optional[str] = None
    manufacturer: Optional[str] = None
    model: Optional[str] = None
    owner: Optional[str] = None
    type: Optional[str] = None


# -----------------------------
# JSON trace parsing
# -----------------------------


def _get_alt_ft(hit: Dict[str, Any]) -> Optional[float]:
    """
    Best-effort altitude in feet from a hit dict produced by extract_hits.
    """
    if "alt_ft" in hit and hit["alt_ft"] is not None:
        try:
            return float(hit["alt_ft"])
        except Exception:
            pass

    ac = hit.get("ac_data") or {}
    alt = ac.get("alt_geom")
    if alt is None or alt == "ground":
        alt = ac.get("alt_baro")
    if alt == "ground":
        return 0.0
    try:
        if alt is not None:
            return float(alt)
    except Exception:
        return None
    return None


def extract_hits(blob: Dict[str, Any]) -> List[List[Dict[str, Any]]]:
    """
    Convert a trace_full JSON blob into a list of segments, where each
    segment is a list of normalized "hit" dicts.

    Uses the JSON format documented by adsblol/globe_history_2024:
    - top-level keys: icao, timestamp, trace
    - each trace row: [dt, lat, lon, alt_ft_or_ground, gs, track, flags, ..., ac_meta, ...]
    """
    base_ts = blob.get("timestamp")
    trace = blob.get("trace") or []

    if not isinstance(trace, list) or base_ts is None:
        return []

    segments: List[List[Dict[str, Any]]] = []
    current: List[Dict[str, Any]] = []

    for row in trace:
        if not isinstance(row, list) or len(row) < 3:
            continue

        try:
            dt_offset = float(row[0])
            lat = float(row[1])
            lon = float(row[2])
        except Exception:
            continue

        ts = float(base_ts) + dt_offset

        alt_field = row[3] if len(row) > 3 else None
        gs = row[4] if len(row) > 4 else None
        track = row[5] if len(row) > 5 else None
        flags = row[6] if len(row) > 6 else 0
        ac_meta = row[8] if len(row) > 8 and isinstance(row[8], dict) else {}

        # Prefer geometric altitude if available in metadata
        alt_geom = None
        if isinstance(ac_meta, dict):
            alt_geom = ac_meta.get("alt_geom")
        alt_ft: Optional[float]
        if alt_geom not in (None, "ground"):
            try:
                alt_ft = float(alt_geom)
            except Exception:
                alt_ft = None
        else:
            if alt_field == "ground":
                alt_ft = 0.0
            else:
                try:
                    alt_ft = float(alt_field) if alt_field is not None else None
                except Exception:
                    alt_ft = None

        hit: Dict[str, Any] = {
            "timestamp": ts,
            "lat": lat,
            "lon": lon,
            "alt_ft": alt_ft,
            "gs": gs,
            "track": track,
            "flags": flags,
            "ac_data": ac_meta,
        }

        start_new_leg = False
        try:
            if isinstance(flags, int) and (flags & 2):
                start_new_leg = True
        except Exception:
            start_new_leg = False

        if start_new_leg and current:
            segments.append(current)
            current = []

        current.append(hit)

    if current:
        segments.append(current)

    return segments


# -----------------------------
# ADSBx Aircraft DB helpers
# -----------------------------


def load_adsbx_acdb(root_dir: str, log_cb=print) -> Any:
    """
    Load the ADSBx basic aircraft DB (JSON.gz) into memory, with full support for:
      • gzip or non-gzip payloads
      • single JSON document
      • newline-delimited JSON (JSONL)
    """
    global _ACDB_CACHE

    os.makedirs(root_dir, exist_ok=True)
    cache_path = os.path.join(root_dir, "basic-ac-db.json.gz")

    def _parse_db_bytes(data: bytes):
        """Parse bytes from ADSBx DB, handling gzip vs plain + JSONL vs JSON."""
        # Try gzip first, fall back to plain text if not gzipped
        try:
            decompressed = gzip.decompress(data)
        except OSError:
            decompressed = data

        text = decompressed.decode("utf-8", errors="ignore").strip()
        if not text:
            return {}

        # First attempt: conventional JSON (dict or list)
        try:
            return json.loads(text)
        except json.JSONDecodeError as e:
            # "Extra data" usually means multiple JSON objects -> JSONL
            if "Extra data" not in str(e):
                raise

        # Second attempt: newline-delimited JSON (one object per line)
        records = []
        for line in text.splitlines():
            line = line.strip()
            if not line or line.startswith("//"):
                continue
            try:
                obj = json.loads(line)
            except Exception:
                continue
            records.append(obj)
        return records

    # Use cached in-memory copy if we already loaded it
    if _ACDB_CACHE is not None:
        return _ACDB_CACHE

    # Try reading from cache on disk
    if os.path.exists(cache_path):
        try:
            with open(cache_path, "rb") as f:
                data = f.read()
            _ACDB_CACHE = _parse_db_bytes(data)
            log_cb("[acdb] Loaded ADSBx aircraft DB from cache")
            return _ACDB_CACHE
        except Exception as e:
            log_cb(f"[acdb] Failed to read cached DB ({e}), re-downloading…")

    # Download fresh copy
    try:
        log_cb("[acdb] Downloading ADSBx aircraft DB…")
        resp = requests.get(AC_DB_URL, timeout=60)
        resp.raise_for_status()
        data = resp.content

        # Save raw bytes to cache (still compressed)
        with open(cache_path, "wb") as f:
            f.write(data)

        _ACDB_CACHE = _parse_db_bytes(data)
        log_cb("[acdb] Downloaded and cached ADSBx aircraft DB")
        return _ACDB_CACHE
    except Exception as e:
        log_cb(f"[acdb] Error downloading ADSBx DB: {e}")
        _ACDB_CACHE = {}
        return _ACDB_CACHE


def find_acdb_record(db: Any, icao_hex: str) -> Optional[dict]:
    """Look up a single ICAO hex record in ADSBx basic aircraft DB."""
    icao_hex = icao_hex.lower()
    if not db:
        return None

    iterable: Iterable[Any]
    if isinstance(db, dict):
        iterable = db.values()
    else:
        iterable = db

    for rec in iterable:
        if not isinstance(rec, dict):
            continue
        icao_field = (
            rec.get("ICAO")
            or rec.get("icao")
            or rec.get("icao24")
            or rec.get("ICAO24")
            or rec.get("hex")
        )
        if not isinstance(icao_field, str):
            continue
        if icao_field.lower() == icao_hex:
            return rec
    return None


def find_acdb_record_by_reg(db: Any, registration: str) -> Optional[Tuple[str, dict]]:
    """Look up ADSBx basic aircraft DB by registration / tail number and return
    (icao_hex, record) if found.
    """
    if not registration:
        return None

    reg_norm = registration.replace("-", "").strip().upper()
    if not reg_norm or not db:
        return None

    iterable: Iterable[Any]
    if isinstance(db, dict):
        iterable = db.values()
    else:
        iterable = db

    for rec in iterable:
        if not isinstance(rec, dict):
            continue

        reg_field = (
            rec.get("REG")
            or rec.get("reg")
            or rec.get("r")
            or rec.get("registration")
            or rec.get("tail")
        )
        if not isinstance(reg_field, str):
            continue

        reg_val = reg_field.replace("-", "").strip().upper()
        if reg_val != reg_norm:
            continue

        icao_field = (
            rec.get("ICAO")
            or rec.get("icao")
            or rec.get("icao24")
            or rec.get("ICAO24")
            or rec.get("hex")
        )
        if not isinstance(icao_field, str):
            continue

        return icao_field.lower(), rec

    return None


def flags_from_dbflags(dbflags: Any) -> Optional[str]:
    """
    Decode ADSBx dbFlags bitfield if available.
    bit 1: military
    bit 2: interesting
    bit 4: PIA
    bit 8: LADD
    """
    try:
        v = int(dbflags)
    except Exception:
        return None
    flags = []
    if v & 1:
        flags.append("Military")
    if v & 2:
        flags.append("Interesting")
    if v & 4:
        flags.append("PIA")
    if v & 8:
        flags.append("LADD")
    return ", ".join(flags) if flags else None


def merge_adsbx_record_into_meta(rec: dict, meta: "AircraftMeta"):
    """Merge an ADSBx DB record into AircraftMeta."""
    if not isinstance(rec, dict):
        return

    reg = rec.get("REG") or rec.get("reg") or rec.get("r")
    if reg and not meta.registration:
        meta.registration = str(reg).strip()

    t = rec.get("ICAOTYPE") or rec.get("icaoType") or rec.get("type")
    if t and not meta.type:
        meta.type = str(t).strip()

    manuf = rec.get("Manufacturer") or rec.get("manufacturer")
    if manuf and not meta.manufacturer:
        meta.manufacturer = str(manuf).strip()

    model = rec.get("Model") or rec.get("model")
    if model and not meta.model:
        meta.model = str(model).strip()

    owner = rec.get("OWNOP") or rec.get("owner") or rec.get("operator")
    if owner and not meta.owner:
        meta.owner = str(owner).strip()

    dbf = rec.get("dbFlags") or rec.get("DBFLAGS")
    if dbf and not meta.flags:
        f = flags_from_dbflags(dbf)
        if f:
            meta.flags = f


# -----------------------------
# External metadata lookups
# -----------------------------


def fetch_opensky_metadata(icao_hex: str) -> OpenSkyMeta:
    """
    Best-effort lookup of registration/manufacturer/model/owner/type from
    OpenSky aircraft database. This is wrapped in try/except by callers.
    """
    url = f"https://opensky-network.org/api/metadata/aircraft/icao/{icao_hex.lower()}"
    resp = requests.get(url, timeout=10)
    if resp.status_code != 200:
        return OpenSkyMeta()
    try:
        data = resp.json()
    except Exception:
        return OpenSkyMeta()

    return OpenSkyMeta(
        registration=data.get("registration"),
        manufacturer=data.get("manufacturername") or data.get("manufacturer"),
        model=data.get("model"),
        owner=data.get("owner"),
        type=data.get("typecode") or data.get("type"),
    )


def fetch_planespotters_photo_and_reg(icao_hex: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Return (photo_url, registration) from planespotters.net public API, if any.
    """
    url = f"https://api.planespotters.net/pub/photos/hex/{icao_hex.lower()}"
    headers = {
        "User-Agent": "adsbx-history-downloader-gui/1.0 (+https://adsbexchange.com)",
        "Accept": "application/json",
    }
    resp = requests.get(url, timeout=10, headers=headers)
    if resp.status_code != 200:
        return None, None

    try:
        data = resp.json()
    except Exception:
        return None, None

    photos = data.get("photos") or []
    if not photos:
        return None, None

    best = photos[0]
    photo_url = None
    reg = None

    def _get_src(val):
        if isinstance(val, dict):
            return val.get("src")
        if isinstance(val, str):
            return val
        return None

    try:
        images = best.get("images") or {}
        photo_url = (
            _get_src(images.get("large"))
            or _get_src(images.get("medium"))
            or _get_src(best.get("thumbnail_large"))
            or _get_src(best.get("thumbnail"))
            or _get_src(best.get("link"))
        )
    except Exception:
        photo_url = None
    try:
        reg = best.get("registration")
    except Exception:
        reg = None
    return photo_url, reg


def apply_type_mapping(meta: AircraftMeta) -> None:
    """Populate type_name from ICAO type code, if possible."""
    if meta.type and not meta.type_name:
        t = meta.type.upper()
        if t in ICAO_TYPE_NAMES:
            meta.type_name = ICAO_TYPE_NAMES[t]


# -----------------------------
# FAA registry lookup (download + local lookup)
# -----------------------------


FAA_REG_URL = "https://registry.faa.gov/database/ReleasableAircraft.zip"
FAA_CACHE_JSON = "faa_registry_cache.json"
RESOURCE_ROOT = os.path.join(os.getcwd(), "resources")
os.makedirs(RESOURCE_ROOT, exist_ok=True)


def load_faa_registry(root_dir: str, log_cb=print) -> Dict[str, Dict[str, Any]]:
    """
    Download and cache the FAA Releasable Aircraft database, returning a dict keyed by N-number.
    """
    os.makedirs(root_dir, exist_ok=True)
    cache_json = os.path.join(root_dir, FAA_CACHE_JSON)
    zip_path = os.path.join(root_dir, "ReleasableAircraft.zip")
    alt_local_zip = os.path.join(os.getcwd(), "faa_db", "ReleasableAircraft.zip")
    alt_local_zip2 = os.path.join(os.getcwd(), "faa_registry", "ReleasableAircraft.zip")

    log_cb(
        f"[faa] FAA paths: cache_json={cache_json}, zip_path={zip_path}, "
        f"alt1_exists={os.path.exists(alt_local_zip)}, alt2_exists={os.path.exists(alt_local_zip2)}"
    )

    # Try cached JSON first
    if os.path.exists(cache_json):
        try:
            with open(cache_json, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                if data:
                    log_cb(f"[faa] Loaded cached FAA registry ({len(data)} records)")
                    return data
                else:
                    log_cb("[faa] Cached FAA registry is empty; will re-parse zip")
        except Exception:
            log_cb("[faa] Failed to read cached FAA registry; re-parsing")

    # Prefer a local FAA DB zip if present (in root_dir or ./faa_db)
    if not os.path.exists(zip_path):
        if os.path.exists(alt_local_zip):
            try:
                shutil.copyfile(alt_local_zip, zip_path)
                log_cb(f"[faa] Using local FAA DB zip from {alt_local_zip}")
            except Exception as e:
                log_cb(f"[faa] Failed to copy local FAA DB zip: {e}")
        elif os.path.exists(alt_local_zip2):
            try:
                shutil.copyfile(alt_local_zip2, zip_path)
                log_cb(f"[faa] Using local FAA DB zip from {alt_local_zip2}")
            except Exception as e:
                log_cb(f"[faa] Failed to copy local FAA DB zip: {e}")
    else:
        log_cb(f"[faa] Using existing FAA DB zip at {zip_path}")

    # If local zip exists in FAA DB folder, prefer it; otherwise attempt download
    download_ok = os.path.exists(zip_path)
    if not download_ok:
        # More robust download with user-agent + retry on connection resets
        session = requests.Session()
        try:
            from urllib3.util.retry import Retry
            from requests.adapters import HTTPAdapter

            retry = Retry(
                total=5,
                backoff_factor=1.5,
                status_forcelist=[429, 500, 502, 503, 504],
                allowed_methods=["GET"],
            )
            adapter = HTTPAdapter(max_retries=retry)
            session.mount("https://", adapter)
            session.mount("http://", adapter)
        except Exception:
            pass

        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) SkyProfile/1.0",
            "Accept": "*/*",
        }

        for attempt in range(5):
            try:
                log_cb(f"[faa] Downloading FAA registry (attempt {attempt+1}/5) from {FAA_REG_URL}")
                resp = session.get(
                    FAA_REG_URL,
                    timeout=(10, 120),
                    stream=True,
                    headers=headers,
                    allow_redirects=True,
                )
                resp.raise_for_status()
                with open(zip_path, "wb") as f:
                    for chunk in resp.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
                download_ok = True
                break
            except Exception as e:
                log_cb(f"[faa] Download failed (attempt {attempt+1}/5): {e}")
                time.sleep(2 * (attempt + 1))

    if not download_ok:
        log_cb("[faa] Using local FAA DB only; no download available.")
        if not os.path.exists(zip_path):
            log_cb("[faa] No local FAA zip found; skipping FAA registry.")
            return {}

    records: Dict[str, Dict[str, Any]] = {}
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            # Prefer MASTER.txt if present
            target = None
            for name in zf.namelist():
                if name.lower().endswith("master.txt"):
                    target = name
                    break
            if not target:
                log_cb("[faa] No MASTER.txt found in FAA archive")
                return {}
            with zf.open(target) as f:
                raw_bytes = f.read()
                try:
                    text = raw_bytes.decode("utf-8")
                except UnicodeDecodeError:
                    # FAA MASTER.txt is often Windows-1252; fall back if needed
                    text = raw_bytes.decode("cp1252", errors="ignore")
            # Parse MASTER with multiple delimiters (pipe, comma) and more tolerant header cleaning
            def parse_master(text: str) -> Dict[str, Dict[str, Any]]:
                for delim in ("|", ","):
                    reader = csv.reader(text.splitlines(), delimiter=delim)
                    rows_all = [r for r in reader if r]
                    if not rows_all:
                        continue

                    def clean_header_cell(cell: str) -> str:
                        cell = (cell or "").strip().strip('"').strip("'")
                        cell = re.sub(r"^[^\w]*", "", cell)  # drop BOM/garbage
                        return cell

                    header = None
                    header_idx = 0
                    for i, row in enumerate(rows_all):
                        cleaned = [clean_header_cell(h) for h in row]
                        upper = [h.upper() for h in cleaned]
                        if "N-NUMBER" in upper or "NNUMBER" in upper:
                            header = cleaned
                            header_idx = i
                            break
                    if header is None:
                        header = [clean_header_cell(h) for h in rows_all[0]]
                        header_idx = 0
                    rows = rows_all[header_idx + 1 :]
                    if not header or not rows:
                        continue

                    idx = {h.upper(): i for i, h in enumerate(header) if h}

                    def get(row, key):
                        i = idx.get(key)
                        if i is None or i >= len(row):
                            return ""
                        return row[i].strip()

                    def norm(s: str) -> str:
                        return re.sub(r"[^A-Z0-9]", "", (s or "").upper())

                    normalized_header = {norm(h): h for h in header if h}

                    def find_value(row, candidates):
                        for cand in candidates:
                            key_norm = norm(cand)
                            h = normalized_header.get(key_norm)
                            if h:
                                val = get(row, h)
                                if val:
                                    return val
                        return ""

                    recs: Dict[str, Dict[str, Any]] = {}
                    for row in rows:
                        nnum = find_value(row, ["N-NUMBER", "NNUMBER"])
                        if not nnum:
                            continue
                        n_prefix = f"N{nnum.upper()}"
                        n_plain = nnum.upper()
                        raw = {h: get(row, h) for h in header}
                        friendly = {
                            "N-Number": n_prefix,
                            "Serial Number": find_value(row, ["SERIAL NUMBER", "SERIALNUMBER"]),
                            "Status": find_value(row, ["STATUS CODE", "STATUS"]),
                            "Manufacturer Name": find_value(row, ["MFR NAME", "MANUFACTURER NAME", "MFRMDLCODE", "MFR MDL CODE"]),
                            "Model": find_value(row, ["MODEL", "MFR MDL CODE", "MFRMDLCODE"]),
                            "Certificate Issue Date": find_value(row, ["CERT ISSUE DATE", "CERTISSUEDATE"]),
                            "Expiration Date": find_value(row, ["EXPIRATION DATE", "EXPIRATION"]),
                            "Type Aircraft": find_value(row, ["TYPE AIRCRAFT", "TYPEAIRCRAFT"]),
                            "Type Engine": find_value(row, ["TYPE ENGINE", "TYPEENGINE"]),
                            "Pending Number Change": find_value(row, ["PENDING NUMBER CHANGE", "PENDINGNUMBERCHANGE"]),
                            "Dealer": find_value(row, ["DEALER"]),
                            "Date Change Authorized": find_value(row, ["DATE CHANGE AUTHORIZED", "DATECHANGEAUTHORIZED"]),
                            "Mode S Code (base 8 / Oct)": find_value(row, ["MODE S CODE", "MODESCODE"]),
                            "Mode S Code (Base 16 / Hex)": find_value(row, ["MODE S CODE HEX", "MODESCODEHEX"]),
                            "MFR Year": find_value(row, ["YEAR MFR", "YEARMFR"]),
                            "Type Registration": find_value(row, ["TYPE REGISTRANT", "TYPEREGISTRANT"]),
                            "Fractional Owner": find_value(row, ["FRACT OWNER", "FRACTOWNER"]),
                            "Registered Owner": find_value(row, ["NAME", "REGISTERED OWNER", "OWNER"]),
                            "Airworthiness Classification": find_value(row, ["AIRWORTHINESS CLASSIFICATION", "AIRWORTHINESSCLASSIFICATION"]),
                            "raw": raw,
                        }
                        recs[n_prefix] = friendly
                        recs[n_plain] = friendly
                    if recs:
                        log_cb(f"[faa] Parsed {len(recs)} FAA records using delimiter '{delim}' from {target}")
                        return recs
                return {}

            records = parse_master(text)
            if not records:
                log_cb("[faa] Parsed 0 FAA records; cache not updated.")
                return {}
    except Exception as e:
        log_cb(f"[faa] Parse failed: {e}")
        return {}

    try:
        with open(cache_json, "w", encoding="utf-8") as f:
            json.dump(records, f)
        log_cb(f"[faa] Cached FAA registry to {cache_json}")
    except Exception:
        pass

    return records


def lookup_faa_registry(n_number: str, db: Dict[str, Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not n_number:
        return None
    n = n_number.strip().upper()
    if not n.startswith("N"):
        return None
    rec = db.get(n)
    if rec:
        return rec
    # Try without N prefix
    rec = db.get(n.lstrip("N"))
    if rec:
        return rec
    return None


# -----------------------------
# Trace fetch from ADSBx
# -----------------------------


def fetch_trace_for_day(
    icao_hex: str,
    day: dt.date,
    session: Optional[requests.Session],
    log_cb=print,
    cache_root: Optional[str] = None,
) -> Optional[str]:
    """
    Fetch trace_full JSON (possibly gzipped) for a given ICAO + date.
    Returns path to cached JSON file, or None if not available.
    """
    y, m, d = day.year, day.month, day.day
    suffix = icao_hex[-2:].lower()
    url = BASE.format(y=y, m=m, d=d, suffix=suffix, hex=icao_hex.lower())

    if cache_root is None:
        cache_root = os.path.join(os.getcwd(), "cache")

    day_dir = os.path.join(cache_root, icao_hex.lower())
    os.makedirs(day_dir, exist_ok=True)
    fn = os.path.join(day_dir, f"{y}-{m:02d}-{d:02d}.json")

    if os.path.exists(fn):
        return fn

    sess = session or requests.Session()
    log_cb(f"[fetch] {day} {icao_hex}")

    try:
        resp = sess.get(url, headers=HEADERS, timeout=30)
    except Exception as e:
        log_cb(f"[fetch] error {day}: {e}")
        return None

    if resp.status_code == 404:
        log_cb(f"[fetch] {day}: 404 (no trace)")
        return None
    if resp.status_code == 429:
        log_cb(f"[fetch] {day}: 429 (rate limited)")
        time.sleep(30)
        return None
    if resp.status_code != 200:
        log_cb(f"[fetch] {day}: HTTP {resp.status_code}")
        return None

    try:
        with open(fn, "wb") as f:
            f.write(resp.content)
    except Exception as e:
        log_cb(f"[fetch] error writing cache: {e}")
        return None

    return fn


def merge_trace_blob_into_meta(blob: Dict[str, Any], meta: AircraftMeta) -> None:
    """
    Merge top-level metadata fields from a trace_full blob into AircraftMeta.
    """
    if not isinstance(blob, dict):
        return

    reg = blob.get("r")
    if reg and not meta.registration:
        meta.registration = str(reg).strip()

    t = blob.get("t")
    if t and not meta.type:
        meta.type = str(t).strip()

    desc = blob.get("desc")
    if desc and not meta.description:
        meta.description = str(desc).strip()

    dbf = blob.get("dbFlags")
    if dbf and not meta.flags:
        f = flags_from_dbflags(dbf)
        if f:
            meta.flags = f


# -----------------------------
# KML: leg-based points / routes / 3D routes
# -----------------------------


def extract_legs_for_kml(
    segments: List[List[Dict[str, Any]]],
    alt_thresh_ft: float = 500.0,
) -> List[Dict[str, Any]]:
    """
    Derive flight legs from raw segments for KML export.

    A leg is defined as: ground -> air -> ground, where "ground" is inferred
    from ADSBx fields plus altitude/groundspeed thresholds. We also treat long
    coverage gaps as implicit landings so multi-leg days like A→B→C→A are split
    even if the transponder is quiet on the ground.
    """
    legs: List[Dict[str, Any]] = []

    # Allow short ground-out periods and long coverage gaps to split legs
    MIN_GROUND_DWELL_S = 90.0
    GAP_FORCE_LANDING_S = 20 * 60.0
    GS_GROUND_MAX = 60.0

    def hit_is_ground(hit: Dict[str, Any]) -> bool:
        """Best-effort on-ground detection using ADSBx metadata and heuristics."""
        alt_ft = _get_alt_ft(hit)
        if alt_ft is not None and 0 <= alt_ft <= alt_thresh_ft:
            return True

        ac = hit.get("ac_data") or {}
        if isinstance(ac, dict):
            ag = str(ac.get("air_ground") or "").lower()
            if ag in ("g", "ground", "0"):
                return True
            gnd = ac.get("gnd") or ac.get("ground") or ac.get("on_ground") or ac.get("onground")
            if isinstance(gnd, bool) and gnd:
                return True

        try:
            gs_val = float(hit.get("gs"))
            if (alt_ft is None or alt_ft <= alt_thresh_ft * 1.5) and gs_val <= GS_GROUND_MAX:
                return True
        except Exception:
            pass

        return False

    def pull_callsign(hit: Dict[str, Any]) -> Optional[str]:
        ac = hit.get("ac_data")
        if not isinstance(ac, dict):
            return None
        for key in ("flight", "call", "callsign", "cs"):
            val = ac.get(key)
            if isinstance(val, str) and val.strip():
                cs_norm = normalize_callsign(val)
                if cs_norm:
                    return cs_norm
        return None

    # Flatten all hits so we can detect legs that span raw segments
    all_hits: List[Tuple[int, float, float, float, Optional[float], Dict[str, Any]]] = []
    for seg_idx, seg in enumerate(segments, 1):
        for hit in seg:
            try:
                lat = float(hit.get("lat"))
                lon = float(hit.get("lon"))
            except (TypeError, ValueError):
                continue
            ts = hit.get("timestamp")
            if not isinstance(ts, (int, float)):
                continue
            alt_ft = _get_alt_ft(hit)
            all_hits.append((seg_idx, float(ts), lat, lon, alt_ft, hit))

    all_hits.sort(key=lambda x: x[1])
    if len(all_hits) < 2:
        return legs

    dep_info: Optional[Dict[str, Any]] = None
    last_seg_for_leg: Optional[int] = None
    track_points: List[Tuple[float, float, float]] = []
    recent_ground_start: Optional[float] = None
    recent_ground_cluster: List[Tuple[float, float]] = []
    last_air_hit: Optional[Tuple[float, float, float, Optional[float]]] = None
    prev_ts: Optional[float] = None
    on_ground = False

    def close_leg(arr_ts: float, arr_lat: float, arr_lon: float):
        nonlocal dep_info, track_points, recent_ground_cluster, recent_ground_start, last_seg_for_leg
        if dep_info is None:
            return
        dep_ts = dep_info["ts"]
        dep_lat = dep_info["lat"]
        dep_lon = dep_info["lon"]
        callsign_for_leg = dep_info["callsign"]
        callsign_history = dep_info.get("callsign_history") or []
        if arr_ts <= dep_ts:
            return

        dep_dt = dt.datetime.fromtimestamp(dep_ts, dt.timezone.utc)
        arr_dt = dt.datetime.fromtimestamp(arr_ts, dt.timezone.utc)
        duration_min = (arr_dt - dep_dt).total_seconds() / 60.0

        legs.append(
            {
                "segment": last_seg_for_leg,
                "dep_ts": dep_ts,
                "arr_ts": arr_ts,
                "dep_dt": dep_dt,
                "arr_dt": arr_dt,
                "dep_lat": dep_lat,
                "dep_lon": dep_lon,
                "arr_lat": arr_lat,
                "arr_lon": arr_lon,
                "duration_min": duration_min,
                "callsign": callsign_for_leg,
                "callsign_history": list(callsign_history),
                "track": list(track_points),
            }
        )

        dep_info = None
        track_points = []
        recent_ground_cluster = []
        recent_ground_start = None
        last_seg_for_leg = None

    for seg_idx, ts, lat, lon, alt_ft, hit in all_hits:
        # Treat long coverage gaps as implicit landings only if the last seen altitude was near ground.
        if dep_info and prev_ts is not None and ts - prev_ts > GAP_FORCE_LANDING_S:
            if last_air_hit:
                arr_ts, arr_lat, arr_lon, arr_alt_ft = last_air_hit
                if arr_alt_ft is not None and arr_alt_ft > alt_thresh_ft * 1.5:
                    # Still high; keep leg open and continue until an actual landing is observed.
                    pass
                else:
                    close_leg(arr_ts, arr_lat, arr_lon)
            else:
                close_leg(ts, lat, lon)

        is_ground = hit_is_ground(hit)

        if is_ground:
            if not on_ground:
                recent_ground_start = ts
                recent_ground_cluster = []
            recent_ground_cluster.append((lat, lon))
        on_ground = is_ground

        alt_m = float(alt_ft) * 0.3048 if alt_ft is not None else 0.0

        if dep_info:
            track_points.append((lon, lat, alt_m))
            last_air_hit = (ts, lat, lon, alt_ft)
            new_cs = pull_callsign(hit)
            if new_cs and new_cs != dep_info["callsign"]:
                dep_info["callsign"] = new_cs
                history = dep_info.get("callsign_history")
                if isinstance(history, list):
                    if not history or history[-1] != new_cs:
                        history.append(new_cs)
                else:
                    dep_info["callsign_history"] = [new_cs]

        # Takeoff: last known ground cluster -> airborne
        if dep_info is None and not is_ground:
            dep_lat = dep_lon = None
            dep_ts = recent_ground_start or ts
            if recent_ground_cluster:
                dep_lat = sum(p[0] for p in recent_ground_cluster) / len(recent_ground_cluster)
                dep_lon = sum(p[1] for p in recent_ground_cluster) / len(recent_ground_cluster)
            else:
                dep_lat = lat
                dep_lon = lon
            callsign_for_leg = pull_callsign(hit)
            dep_info = {
                "ts": dep_ts,
                "lat": dep_lat,
                "lon": dep_lon,
                "callsign": callsign_for_leg,
                "callsign_history": [callsign_for_leg] if callsign_for_leg else [],
            }
            last_seg_for_leg = seg_idx
            track_points = [(dep_lon, dep_lat, alt_m)]
            last_air_hit = (dep_ts, dep_lat, dep_lon, alt_ft)

        # Landing: sustained ground after being airborne
        if dep_info and is_ground and recent_ground_start is not None:
            if ts - recent_ground_start >= MIN_GROUND_DWELL_S:
                if recent_ground_cluster:
                    arr_lat = sum(p[0] for p in recent_ground_cluster) / len(recent_ground_cluster)
                    arr_lon = sum(p[1] for p in recent_ground_cluster) / len(recent_ground_cluster)
                else:
                    arr_lat, arr_lon = lat, lon
                close_leg(recent_ground_start, arr_lat, arr_lon)

        prev_ts = ts

    return legs


def _describe_leg_for_kml(
    leg: Dict[str, Any], meta: Dict[str, Any]
) -> Tuple[str, Dict[str, str], str]:
    """
    Build a human-readable description and ExtendedData dict for a leg.

    Returns (title, extended_data, description_text).
    """
    dep_dt: dt.datetime = leg["dep_dt"]
    arr_dt: dt.datetime = leg["arr_dt"]
    duration_min: float = leg["duration_min"]
    callsign: Optional[str] = leg.get("callsign")
    dep_ap = leg.get("dep_airport")
    arr_ap = leg.get("arr_airport")

    def ap_label(ap: Optional[Dict[str, Any]]) -> Optional[str]:
        if not ap:
            return None
        codes = [c for c in (ap.get("icao_code"), ap.get("iata_code")) if c]
        base = " / ".join(codes) if codes else ap.get("airport_id") or None
        name = ap.get("name") or ""
        city = ap.get("city") or ""
        country = ap.get("iso_country") or ""
        parts = [p for p in [base, name] if p]
        loc = ", ".join([p for p in [city, country] if p])
        if loc:
            parts.append(loc)
        return " – ".join(parts) if parts else None

    title = dep_dt.strftime("%Y-%m-%d")
    if callsign:
        title += f" – {callsign}"
    dep_label = ap_label(dep_ap)
    arr_label = ap_label(arr_ap)
    if dep_label or arr_label:
        title += f" ({dep_label or 'Unknown'} → {arr_label or 'Unknown'})"

    desc_lines = [
        f"Departure (UTC): {dep_dt.isoformat()}Z",
        f"Arrival (UTC):   {arr_dt.isoformat()}Z",
        f"Duration (min):  {duration_min:.1f}",
    ]
    if dep_label:
        desc_lines.append(f"Departure airport: {dep_label}")
    if arr_label:
        desc_lines.append(f"Arrival airport:   {arr_label}")
    if callsign:
        desc_lines.append(f"Callsign: {callsign}")

    # Aircraft meta – skip obvious location/altitude fields; keep identity info
    for key in (
        "registration",
        "type",
        "type_name",
        "owner",
        "manufacturer",
        "model",
        "flags",
        "description",
    ):
        val = meta.get(key)
        if val:
            desc_lines.append(f"{key.capitalize()}: {val}")

    ext: Dict[str, str] = {
        "dep_time_utc": dep_dt.isoformat() + "Z",
        "arr_time_utc": arr_dt.isoformat() + "Z",
        "duration_min": f"{duration_min:.1f}",
    }
    if callsign:
        ext["callsign"] = callsign
    for key in ("registration", "type", "type_name", "owner", "manufacturer", "model", "flags"):
        val = meta.get(key)
        if val:
            ext[f"meta_{key}"] = str(val)

    desc = "\n".join(desc_lines)
    return title, ext, desc


def build_kml_points(
    legs: List[Dict[str, Any]],
    hex_code: str,
    meta: Dict[str, Any],
    out_path: str,
) -> None:
    """
    KML containing points only, one per leg, at the arrival (on-ground) location.

    Each point contains summary information about that leg in its description
    and ExtendedData.
    """
    ensure_dir_for_file(out_path)
    kml = simplekml.Kml()
    root_name = f"SkyProfile {hex_code.upper()} – Points"
    root = kml.newfolder(name=root_name)

    meta_lines = [f"ICAO: {hex_code.upper()}"]
    for k in ("registration", "type", "type_name", "owner", "description"):
        if meta.get(k):
            meta_lines.append(f"{k.capitalize()}: {meta[k]}")
    calls_meta = meta.get("callsigns")
    if calls_meta:
        meta_lines.append(f"Callsigns: {', '.join(calls_meta)}")
    root.description = "\n".join(meta_lines)

    for idx, leg in enumerate(legs, 1):
        arr_lat = float(leg["arr_lat"])
        arr_lon = float(leg["arr_lon"])
        title, ext, desc = _describe_leg_for_kml(leg, meta)

        p = root.newpoint(
            name=f"Leg {idx}: {title}",
            coords=[(arr_lon, arr_lat)],
        )
        p.altitudemode = simplekml.AltitudeMode.clamptoground
        p.description = desc
        for k, v in ext.items():
            try:
                p.extendeddata.simplenode(k, v)
            except Exception:
                pass

    if not legs:
        root.newpoint(name="No legs detected", coords=[])

    kml.save(out_path)


def build_kml_routes_2d(
    legs: List[Dict[str, Any]],
    hex_code: str,
    meta: Dict[str, Any],
    out_path: str,
    max_points_per_leg: int = 1000,
) -> None:
    """
    KML containing 2D routes: along-track polylines per leg (no altitude).
    """
    ensure_dir_for_file(out_path)
    kml = simplekml.Kml()
    root_name = f"SkyProfile {hex_code.upper()} – Routes 2D"
    root = kml.newfolder(name=root_name)

    meta_lines = [f"ICAO: {hex_code.upper()}"]
    for k in ("registration", "type", "type_name", "owner", "description"):
        if meta.get(k):
            meta_lines.append(f"{k.capitalize()}: {meta[k]}")
    calls_meta = meta.get("callsigns")
    if calls_meta:
        meta_lines.append(f"Callsigns: {', '.join(calls_meta)}")
    root.description = "\n".join(meta_lines)

    def decimate_track(
        track: List[Tuple[float, float, float]], max_points: int
    ) -> List[Tuple[float, float, float]]:
        n = len(track)
        if n <= max_points or max_points <= 0:
            return track
        step = max(1, n // max_points)
        return track[::step]

    for idx, leg in enumerate(legs, 1):
        dep_lat = float(leg["dep_lat"])
        dep_lon = float(leg["dep_lon"])
        arr_lat = float(leg["arr_lat"])
        arr_lon = float(leg["arr_lon"])
        track = leg.get("track") or []
        coords_src = track if len(track) >= 2 else [(dep_lon, dep_lat, 0.0), (arr_lon, arr_lat, 0.0)]
        coords_dec = decimate_track(coords_src, max_points_per_leg)
        coords_2d = [(lon, lat, 0.0) for (lon, lat, _alt) in coords_dec]
        title, ext, desc = _describe_leg_for_kml(leg, meta)

        ls = root.newlinestring(
            name=f"Leg {idx}: {title}",
            coords=coords_2d,
        )
        ls.altitudemode = simplekml.AltitudeMode.clamptoground
        ls.extrude = 0
        ls.tessellate = 1
        ls.description = desc
        for k, v in ext.items():
            try:
                ls.extendeddata.simplenode(k, v)
            except Exception:
                pass

    if not legs:
        root.newpoint(name="No legs detected", coords=[])

    kml.save(out_path)


def build_kml_routes_3d(
    legs: List[Dict[str, Any]],
    hex_code: str,
    meta: Dict[str, Any],
    out_path: str,
    max_points_per_leg: int = 1000,
) -> None:
    """
    KML containing 3D routes: along-track polylines with altitude per leg.

    Track is decimated to at most max_points_per_leg points to avoid gigantic
    KML files.
    """
    ensure_dir_for_file(out_path)
    kml = simplekml.Kml()
    root_name = f"SkyProfile {hex_code.upper()} – Routes 3D"
    root = kml.newfolder(name=root_name)

    meta_lines = [f"ICAO: {hex_code.upper()}"]
    for k in ("registration", "type", "type_name", "owner", "description"):
        if meta.get(k):
            meta_lines.append(f"{k.capitalize()}: {meta[k]}")
    calls_meta = meta.get("callsigns")
    if calls_meta:
        meta_lines.append(f"Callsigns: {', '.join(calls_meta)}")
    root.description = "\n".join(meta_lines)

    def decimate_track(
        track: List[Tuple[float, float, float]], max_points: int
    ) -> List[Tuple[float, float, float]]:
        n = len(track)
        if n <= max_points or max_points <= 0:
            return track
        step = max(1, n // max_points)
        return track[::step]

    def altitude_to_color_m(alt_m: float, min_ft: float = 0.0, max_ft: float = 45000.0) -> str:
        """
        Map altitude (meters) to color. Currently returns white for all legs.
        """
        return "ffffffff"

    for idx, leg in enumerate(legs, 1):
        track = leg.get("track") or []
        if len(track) < 2:
            continue
        coords = decimate_track(track, max_points_per_leg)
        title, ext, desc = _describe_leg_for_kml(leg, meta)

        # Use max altitude in leg to choose color
        max_alt_m = max((p[2] for p in coords), default=0.0)
        color_hex = altitude_to_color_m(max_alt_m)

        ls = root.newlinestring(
            name=f"Leg {idx}: {title}",
            coords=coords,
        )
        ls.altitudemode = simplekml.AltitudeMode.absolute
        ls.extrude = 0  # free-floating
        ls.tessellate = 0
        ls.style.linestyle.color = color_hex
        ls.style.linestyle.width = 3
        ls.description = desc
        for k, v in ext.items():
            try:
                ls.extendeddata.simplenode(k, v)
            except Exception:
                pass

    if not legs:
        root.newpoint(name="No legs detected", coords=[])

    kml.save(out_path)


# -----------------------------
# CSV / JSON builders
# -----------------------------


def build_csv(
    segments: List[List[Dict[str, Any]]],
    hex_code: str,
    meta: Dict[str, Any],
    out_path: str,
) -> None:
    """
    Build a consolidated CSV across all segments.

    Columns:
        hex, segment, idx, timestamp_utc, lat, lon, alt_ft, gs, track,
        plus one column per key in ac_data (flattened).
    """
    ensure_dir_for_file(out_path)

    # Gather all ac_data keys
    ac_keys: Set[str] = set()
    for seg in segments:
        for hit in seg:
            ac = hit.get("ac_data")
            if isinstance(ac, dict):
                ac_keys.update(ac.keys())

    base_cols = [
        "hex",
        "segment",
        "idx",
        "timestamp_utc",
        "lat",
        "lon",
        "alt_ft",
        "gs",
        "track",
    ]
    ac_cols = sorted(ac_keys)
    cols = base_cols + ac_cols

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=cols)
        writer.writeheader()

        for seg_idx, seg in enumerate(segments, 1):
            for idx, hit in enumerate(seg, 1):
                row = {
                    "hex": hex_code.upper(),
                    "segment": seg_idx,
                    "idx": idx,
                    "timestamp_utc": dt.datetime.fromtimestamp(
                        float(hit.get("timestamp", 0.0)), dt.timezone.utc
                    ).isoformat()
                    + "Z",
                    "lat": hit.get("lat"),
                    "lon": hit.get("lon"),
                    "alt_ft": _get_alt_ft(hit),
                    "gs": hit.get("gs"),
                    "track": hit.get("track"),
                }
                ac = hit.get("ac_data") or {}
                if not isinstance(ac, dict):
                    ac = {}
                for k in ac_cols:
                    v = ac.get(k)
                    if isinstance(v, (dict, list)):
                        v = json.dumps(v, ensure_ascii=False)
                    row[k] = v
                writer.writerow(row)


def build_json(raw_blobs: List[Dict[str, Any]], out_path: str) -> None:
    """
    Save a merged collection of the raw ADSBx trace_full JSON blobs (no additional parsing).
    """
    ensure_dir_for_file(out_path)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(raw_blobs, f, ensure_ascii=False, indent=2, default=str)


# -----------------------------
# Excel summary (with airports, routes, countries, callsigns, calendar)
# -----------------------------


def load_airport_db_for_summary(root_dir: str, log_cb=print) -> Dict[str, Any]:
    """
    Load an airport database for Excel summaries.

    Expected formats (any one of these, placed in `root_dir`):

    1) airports.json
       - A JSON list of airport dicts with keys like:
         id / ident, name, city, iso_country, type, latitude_deg, longitude_deg,
         icao_code, iata_code, gps_code, etc.

    2) airports.csv
       - A CSV with columns similar to the OpenFlights schema:
         id, ident, type, name, latitude_deg, longitude_deg, iso_country,
         municipality (city), iata_code, gps_code, etc.

    If no usable file is found, returns {"airports": []} and the code will still
    generate Flights / DayOfWeek / DayOfMonth / Hours / Callsigns,
    but Airports / Routes / Countries sheets will be empty.
    """
    import csv

    os.makedirs(root_dir, exist_ok=True)

    json_path = os.path.join(root_dir, "airports.json")
    csv_path = os.path.join(root_dir, "airports.csv")

    airports: List[Dict[str, Any]] = []

    def norm_airport(rec: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        # Normalize one airport record to the shape we want for summaries.
        try:
            lat = (
                rec.get("lat")
                or rec.get("latitude_deg")
                or rec.get("latitude")
                or rec.get("LATITUDE_DEG")
            )
            lon = (
                rec.get("lon")
                or rec.get("longitude_deg")
                or rec.get("longitude")
                or rec.get("LONGITUDE_DEG")
            )
            if lat is None or lon is None:
                return None
            lat = float(lat)
            lon = float(lon)
        except Exception:
            return None

        ident = (
            rec.get("airport_id")
            or rec.get("ident")
            or rec.get("id")
            or rec.get("gps_code")
            or rec.get("icao_code")
        )
        if not ident:
            return None

        icao_code = rec.get("icao_code") or rec.get("gps_code")
        iata_code = rec.get("iata_code") or rec.get("iata")
        name = rec.get("name")
        city = rec.get("city") or rec.get("municipality")
        iso_country = rec.get("iso_country") or rec.get("country")
        atype = rec.get("type")

        return {
            "airport_id": str(ident),
            "icao_code": icao_code or "",
            "iata_code": iata_code or "",
            "name": name or "",
            "city": city or "",
            "iso_country": iso_country or "",
            "type": atype or "",
            "lat": lat,
            "lon": lon,
        }

    # Prefer JSON if present
    if os.path.exists(json_path):
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                iterable = data.get("airports") or data.get("data") or []
            else:
                iterable = data
            for rec in iterable:
                if not isinstance(rec, dict):
                    continue
                out = norm_airport(rec)
                if out:
                    airports.append(out)
            log_cb(f"[airports] Loaded {len(airports)} airports from {json_path}")
            return {"airports": airports}
        except Exception as e:
            log_cb(f"[airports] Failed to load {json_path}: {e}")

    # Try CSV
    if os.path.exists(csv_path):
        try:
            with open(csv_path, "r", encoding="utf-8", newline="") as f:
                reader = csv.DictReader(f)
                for rec in reader:
                    out = norm_airport(rec)
                    if out:
                        airports.append(out)
            log_cb(f"[airports] Loaded {len(airports)} airports from {csv_path}")
            return {"airports": airports}
        except Exception as e:
            log_cb(f"[airports] Failed to load {csv_path}: {e}")

    # Attempt to download a public airport DB (OurAirports) if nothing local
    ourairports_url = "https://ourairports.com/data/airports.csv"
    try:
        log_cb(f"[airports] Downloading OurAirports DB from {ourairports_url}")
        resp = requests.get(ourairports_url, timeout=30)
        resp.raise_for_status()
        with open(csv_path, "wb") as f:
            f.write(resp.content)
        with open(csv_path, "r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            for rec in reader:
                out = norm_airport(rec)
                if out:
                    airports.append(out)
        log_cb(f"[airports] Loaded {len(airports)} airports from downloaded OurAirports DB")
        return {"airports": airports}
    except Exception as e:
        log_cb(f"[airports] No usable airport DB found ({e}); Airports/Routes/Countries sheets will be limited.")

    return {"airports": []}


def normalize_airports_for_matching(airport_db: Any) -> List[Dict[str, Any]]:
    airports_raw: List[Dict[str, Any]] = []
    if isinstance(airport_db, dict):
        airports_raw = airport_db.get("airports") or []
    elif isinstance(airport_db, list):
        airports_raw = airport_db

    airports: List[Dict[str, Any]] = []
    for rec in airports_raw:
        try:
            lat = rec.get("lat")
            lon = rec.get("lon")
            if lat is None or lon is None:
                continue
            lat = float(lat)
            lon = float(lon)
        except Exception:
            continue
        airports.append(
            {
                "airport_id": str(rec.get("airport_id") or rec.get("ident") or rec.get("id") or ""),
                "icao_code": rec.get("icao_code") or "",
                "iata_code": rec.get("iata_code") or "",
                "name": rec.get("name") or "",
                "city": rec.get("city") or "",
                "iso_country": rec.get("iso_country") or "",
                "type": rec.get("type") or "",
                "lat": lat,
                "lon": lon,
            }
        )
    return airports


def augment_legs_with_airports(
    legs: List[Dict[str, Any]],
    airports: List[Dict[str, Any]],
    max_km: float = 250.0,
) -> Tuple[
    List[Dict[str, Any]],
    Dict[str, Dict[str, Any]],
    Dict[Tuple[str, str], Dict[str, Any]],
    Set[str],
]:
    """
    Attach nearest-airport info to legs (dep/arr), returning augmented legs plus airport/route stats.
    """

    def haversine_km(lat1, lon1, lat2, lon2) -> float:
        R = 6371.0
        phi1 = math.radians(lat1)
        phi2 = math.radians(lat2)
        dphi = math.radians(lat2 - lat1)
        dlambda = math.radians(lon2 - lon1)
        a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(
            dlambda / 2
        ) ** 2
        c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
        return R * c

    def nearest_airport(lat: float, lon: float) -> Optional[Dict[str, Any]]:
        if not airports:
            return None
        best = None
        best_d = None
        for ap in airports:
            d = haversine_km(lat, lon, ap["lat"], ap["lon"])
            if best_d is None or d < best_d:
                best_d = d
                best = ap
        if best_d is None or best_d > max_km:
            return None
        return best

    def ensure_airport_record(
        ap: Optional[Dict[str, Any]], lat: float, lon: float
    ) -> Dict[str, Any]:
        if ap:
            return ap
        aid = f"LL_{lat:.3f}_{lon:.3f}"
        return {
            "airport_id": aid,
            "icao_code": "",
            "iata_code": "",
            "name": "",
            "city": "",
            "iso_country": "",
            "type": "unknown",
            "lat": lat,
            "lon": lon,
        }

    airport_stats: Dict[str, Dict[str, Any]] = {}
    route_stats: Dict[Tuple[str, str], Dict[str, Any]] = {}
    countries_set: Set[str] = set()
    leg_augmented: List[Dict[str, Any]] = []

    for i, leg in enumerate(legs, 1):
        dep_lat = float(leg["dep_lat"])
        dep_lon = float(leg["dep_lon"])
        arr_lat = float(leg["arr_lat"])
        arr_lon = float(leg["arr_lon"])
        dep_dt_raw: dt.datetime = leg["dep_dt"]
        arr_dt_raw: dt.datetime = leg["arr_dt"]
        dep_dt = dep_dt_raw.replace(tzinfo=None) if isinstance(dep_dt_raw, dt.datetime) and dep_dt_raw.tzinfo else dep_dt_raw
        arr_dt = arr_dt_raw.replace(tzinfo=None) if isinstance(arr_dt_raw, dt.datetime) and arr_dt_raw.tzinfo else arr_dt_raw
        callsign = leg.get("callsign") or ""
        callsign_history = leg.get("callsign_history")

        dep_airport_raw = nearest_airport(dep_lat, dep_lon)
        arr_airport_raw = nearest_airport(arr_lat, arr_lon)

        dep_airport = ensure_airport_record(dep_airport_raw, dep_lat, dep_lon)
        arr_airport = ensure_airport_record(arr_airport_raw, arr_lat, arr_lon)

        dep_id = dep_airport["airport_id"]
        arr_id = arr_airport["airport_id"]

        for ap, role in ((dep_airport, "dep_count"), (arr_airport, "arr_count")):
            aid = ap["airport_id"]
            st = airport_stats.setdefault(
                aid,
                {
                    "airport": ap,
                    "dep_count": 0,
                    "arr_count": 0,
                    "total_visits": 0,
                },
            )
            st[role] += 1
            st["total_visits"] += 1
            country_name = iso_country_name(ap.get("iso_country"))
            if country_name:
                countries_set.add(country_name)

        if dep_id != arr_id:
            key = (dep_id, arr_id)
            rst = route_stats.setdefault(
                key,
                {
                    "dep_airport": dep_airport,
                    "arr_airport": arr_airport,
                    "count": 0,
                    "dates": set(),
                },
            )
            rst["count"] += 1
            rst["dates"].add(dep_dt.date().isoformat())

        leg_augmented.append(
            {
                "leg_index": i,
                "segment": leg.get("segment"),
                "dep_dt": dep_dt,
                "arr_dt": arr_dt,
                "duration_min": float(leg.get("duration_min", 0.0)),
                "dep_lat": dep_lat,
                "dep_lon": dep_lon,
                "arr_lat": arr_lat,
                "arr_lon": arr_lon,
                "dep_airport_id": dep_id,
                "arr_airport_id": arr_id,
                "dep_airport": dep_airport,
                "arr_airport": arr_airport,
                "track": leg.get("track") or [],
                "callsign": callsign,
                "callsign_history": callsign_history,
            }
        )

    return leg_augmented, airport_stats, route_stats, countries_set


def build_summary_excel(
    segments: List[List[Dict[str, Any]]],
    hex_code: str,
    meta_obj: AircraftMeta,
    start_date: dt.date,
    end_date: dt.date,
    out_path: str,
    airport_db: Any,
) -> None:
    """
    Build an Excel workbook with multiple sheets:

      - Summary
      - Flights
      - Airports
      - TopAirports
      - Routes
      - TopRoutes
      - DayOfWeek
      - DayOfMonth
      - Hours
      - Countries
      - Callsigns
    """
    ensure_dir_for_file(out_path)
    wb = Workbook()

    def add_chart_safe(ws, chart, anchor: str):
        # Insert chart normally (legacy behavior). If this raises, let it surface so we can fix it.
        ws.add_chart(chart, anchor)
        # -------- Callsigns seen anywhere in the data (not just legs) ----------
    callsigns_seen_anywhere: Set[str] = set()
    for seg in segments:
        for hit in seg:
            ac = hit.get("ac_data")
            if not isinstance(ac, dict):
                continue
            lower_map = {k.lower(): k for k in ac.keys()}
            for key in ("call", "callsign", "cs", "flight"):
                real = lower_map.get(key)
                if not real:
                    continue
                v = ac.get(real)
                if isinstance(v, str):
                    cs_norm = normalize_callsign(v)
                    if cs_norm:
                        callsigns_seen_anywhere.add(cs_norm)


    # -------- legs from segments ----------
    legs = extract_legs_for_kml(segments)

    # -------- Airport DB normalisation ----------
    # -------- legs from segments ----------
    legs = extract_legs_for_kml(segments)

    airports = normalize_airports_for_matching(airport_db)
    leg_augmented, airport_stats, route_stats, countries_set = augment_legs_with_airports(
        legs, airports
    )

    # ---------- Summary sheet ----------

    ws_summary = wb.active
    ws_summary.title = "Summary"

    row = 1

    def set_row(label: str, value: Any):
        nonlocal row
        ws_summary[f"A{row}"] = label
        ws_summary[f"B{row}"] = value if value is not None else ""
        row += 1

    set_row("ICAO Hex", hex_code.upper())
    set_row("Registration", meta_obj.registration)
    set_row("Type (ICAO)", meta_obj.type)
    set_row("Type name", meta_obj.type_name)
    set_row("Owner", meta_obj.owner)
    set_row("Manufacturer", meta_obj.manufacturer)
    set_row("Model", meta_obj.model)
    set_row("Flags", meta_obj.flags)
    set_row("Date range", f"{start_date.isoformat()} – {end_date.isoformat()}")
    set_row("Total legs", len(legs))
    set_row("Countries visited", ", ".join(sorted(countries_set)) if countries_set else "")

    faa_data = getattr(meta_obj, "faa_data", None) or {}
    if faa_data:
        set_row("FAA N-Number", clean_cell_value(faa_data.get("N-Number") or meta_obj.registration or ""))
        set_row("FAA Status", clean_cell_value(faa_data.get("Status")))
        set_row("FAA Owner", clean_cell_value(faa_data.get("Registered Owner")))
        set_row("FAA Manufacturer", clean_cell_value(faa_data.get("Manufacturer Name")))
        set_row("FAA Model", clean_cell_value(faa_data.get("Model")))
        set_row("FAA Serial Number", clean_cell_value(faa_data.get("Serial Number")))
        set_row("FAA Cert Issue Date", clean_cell_value(faa_data.get("Certificate Issue Date")))
        set_row("FAA Expiration Date", clean_cell_value(faa_data.get("Expiration Date")))
        set_row("FAA Type Aircraft", clean_cell_value(faa_data.get("Type Aircraft")))
        set_row("FAA Type Engine", clean_cell_value(faa_data.get("Type Engine")))
        set_row("FAA Type Registration", clean_cell_value(faa_data.get("Type Registration")))
        set_row("FAA Airworthiness", clean_cell_value(faa_data.get("Airworthiness Classification")))
        set_row("FAA Mode S (Oct)", clean_cell_value(faa_data.get("Mode S Code (base 8 / Oct)")))
        set_row("FAA Mode S (Hex)", clean_cell_value(faa_data.get("Mode S Code (Base 16 / Hex)")))
        set_row("FAA MFR Year", clean_cell_value(faa_data.get("MFR Year")))
        set_row("FAA Pending Number Change", clean_cell_value(faa_data.get("Pending Number Change")))
        set_row("FAA Date Change Authorized", clean_cell_value(faa_data.get("Date Change Authorized")))
        set_row("FAA Dealer", clean_cell_value(faa_data.get("Dealer")))
        set_row("FAA Fractional Owner", clean_cell_value(faa_data.get("Fractional Owner")))

    for col in (1, 2):
        ws_summary.column_dimensions[get_column_letter(col)].width = 24

    # Registered owner details (FAA)
    faa_data = getattr(meta_obj, "faa_data", None) or {}
    # FAA data tab
    ws_faa = wb.create_sheet("FAAData")
    ws_faa.append(["Field", "Value"])
    if faa_data:
        raw = faa_data.get("raw") if isinstance(faa_data, dict) else None
        for key, val in sorted(faa_data.items()):
            if key == "raw":
                continue
            ws_faa.append([clean_cell_value(key), clean_cell_value(val)])
        if isinstance(raw, dict):
            ws_faa.append([])
            ws_faa.append(["Raw FAA fields", ""])
            for k, v in sorted(raw.items()):
                ws_faa.append([clean_cell_value(k), clean_cell_value(v)])
    else:
        ws_faa.append(["No FAA data available for this registration.", ""])

    for col in (1, 2):
        ws_faa.column_dimensions[get_column_letter(col)].width = 32

    # ---------- Aggregation dicts ----------
    dow_counts: Dict[int, int] = collections.defaultdict(int)
    dom_counts: Dict[int, int] = collections.defaultdict(int)
    hour_takeoffs: Dict[int, int] = collections.defaultdict(int)
    hour_landings: Dict[int, int] = collections.defaultdict(int)
    callsign_counts: Dict[str, int] = collections.defaultdict(int)

    # ---------- Flights sheet ----------
    ws_flights = wb.create_sheet("Flights")
    flight_headers = [
        "segment",
        "dep_time_utc",
        "arr_time_utc",
        "duration_min",
        "dep_airport_id",
        "arr_airport_id",
        "dep_lat",
        "dep_lon",
        "arr_lat",
        "arr_lon",
        "dep_dow_index",
        "dep_dow_name",
        "dep_dom",
        "dep_hour_utc",
        "arr_hour_utc",
        "callsign",
        "callsign_history",
    ]
    ws_flights.append(flight_headers)

    for leg in leg_augmented:
        dep_dt = leg["dep_dt"]
        arr_dt = leg["arr_dt"]
        dow_index = dep_dt.weekday()
        dow_name = dep_dt.strftime("%A")
        dom = dep_dt.day
        dep_hour = dep_dt.hour
        arr_hour = arr_dt.hour
        callsign = leg.get("callsign") or ""
        callsign_history = leg.get("callsign_history") or []
        if not isinstance(callsign_history, list):
            callsign_history = [callsign_history] if callsign_history else []
        callsign_history_str = ", ".join([c for c in callsign_history if c])

        dow_counts[dow_index] += 1
        dom_counts[dom] += 1
        hour_takeoffs[dep_hour] += 1
        hour_landings[arr_hour] += 1
        if callsign:
            callsign_counts[callsign] += 1

        row = [
            leg["segment"],
            dep_dt,
            arr_dt,
            leg["duration_min"],
            leg["dep_airport_id"],
            leg["arr_airport_id"],
            leg["dep_lat"],
            leg["dep_lon"],
            leg["arr_lat"],
            leg["arr_lon"],
            dow_index,
            dow_name,
            dom,
            dep_hour,
            arr_hour,
            callsign,
            callsign_history_str,
        ]
        ws_flights.append(row)
    # Ensure any callsign seen in the raw data appears at least once
    for cs in callsigns_seen_anywhere:
        if cs not in callsign_counts:
            callsign_counts[cs] = 1

    
    for col in range(1, len(flight_headers) + 1):
        ws_flights.column_dimensions[get_column_letter(col)].width = 16
    # Make timestamp columns wider to show full datetime
    ts_cols = ["dep_time_utc", "arr_time_utc"]
    for idx, header in enumerate(flight_headers, start=1):
        if header in ts_cols:
            ws_flights.column_dimensions[get_column_letter(idx)].width = 24

    def airport_display_label(ap: Dict[str, Any]) -> str:
        label = ap["airport_id"]
        codes = []
        if ap["icao_code"]:
            codes.append(ap["icao_code"])
        if ap["iata_code"]:
            codes.append(ap["iata_code"])
        if codes:
            label = " / ".join(codes) + " – " + (ap["name"] or "")
        if ap["city"] or ap["iso_country"]:
            city_country = ", ".join([p for p in [ap["city"], ap["iso_country"]] if p])
            if city_country:
                label += " – " + city_country
        return label

    # ---------- Airports (top) sheet ----------
    ws_airports_top = wb.create_sheet("Airports")
    ws_airports_top.append(
        [
            "airport_id",
            "airport",
            "icao_code",
            "iata_code",
            "name",
            "city",
            "iso_country",
            "type",
            "lat",
            "lon",
            "dep_count",
            "arr_count",
            "total_visits",
        ]
    )

    top_airports = sorted(
        airport_stats.items(), key=lambda kv: (-kv[1]["total_visits"], kv[0])
    )[:20]

    for aid, st in top_airports:
        ap = st["airport"]
        label = airport_display_label(ap)
        ws_airports_top.append(
            [
                ap["airport_id"],
                label,
                ap["icao_code"],
                ap["iata_code"],
                ap["name"],
                ap["city"],
                ap["iso_country"],
                ap["type"],
                ap["lat"],
                ap["lon"],
                st["dep_count"],
                st["arr_count"],
                st["total_visits"],
            ]
        )

    for col in range(1, 14):
        ws_airports_top.column_dimensions[get_column_letter(col)].width = 24

    def apply_chart_color(chart_obj, rgb_hex: str = "1f77b4", colors: Optional[List[str]] = None):
        """Force bar colors for readability."""
        try:
            series_list = getattr(chart_obj, "series", []) or []
            for idx, s in enumerate(series_list):
                color = rgb_hex
                if colors and idx < len(colors):
                    color = colors[idx]
                if hasattr(s, "graphicalProperties"):
                    gp = s.graphicalProperties
                    gp.solidFill = color
        except Exception:
            pass

    def add_value_labels(chart_obj):
        """Show category + value on bars, hide series name."""
        try:
            d = DataLabelList()
            d.showVal = True
            d.showSerName = False
            d.showCatName = True
            d.separator = ", "
            d.showLegendKey = False
            chart_obj.dataLabels = d
            # Also blank out any series titles
            if getattr(chart_obj, "series", None):
                for s in chart_obj.series:
                    s.title = None
        except Exception:
            pass

    def configure_chart_axes(chart_obj, x_title: str, y_title: str, max_val: Optional[int] = None):
        """Set axis titles and integer scaling/gridlines."""
        try:
            chart_obj.x_axis.title = x_title
            chart_obj.y_axis.title = y_title
            chart_obj.x_axis.tickLblPos = "low"
            chart_obj.y_axis.number_format = "0"
            chart_obj.y_axis.majorUnit = 1
            chart_obj.y_axis.crosses = "min"
            chart_obj.y_axis.majorGridlines = ChartLines()
            if max_val is not None and max_val > 0:
                chart_obj.y_axis.scaling.min = 0
                chart_obj.y_axis.scaling.max = max_val + 1
        except Exception:
            pass

    # Precompute max values for axis scaling
    max_air_visits = max((st["total_visits"] for _aid, st in top_airports), default=0)

    if ws_airports_top.max_row > 1:
        chart = BarChart()
        chart.title = "Airports by Number of Visits"
        chart.legend = None
        configure_chart_axes(chart, x_title="Airport", y_title="Visits", max_val=max_air_visits)
        data = Reference(
            ws_airports_top,
            min_col=13,
            min_row=2,
            max_row=ws_airports_top.max_row,
        )
        cats = Reference(
            ws_airports_top,
            min_col=2,
            min_row=2,
            max_row=ws_airports_top.max_row,
        )
        chart.add_data(data, titles_from_data=False)
        # Rename series to "flights" to avoid default "Series1"
        try:
            for s in chart.series:
                s.title = "flights"
        except Exception:
            pass
        chart.set_categories(cats)
        chart.height = 16
        chart.width = 32
        add_value_labels(chart)
        apply_chart_color(chart)
        add_chart_safe(ws_airports_top, chart, "N2")
        ws_airports_top["N20"] = "Airports by total visits"

    def airport_label(ap: Optional[Dict[str, Any]]) -> str:
        if not ap:
            return "Unknown"
        return airport_display_label(ap)

    # ---------- Routes (top) sheet ----------
    ws_routes_top = wb.create_sheet("Routes")
    ws_routes_top.append(["route_label", "count"])

    top_routes = sorted(
        route_stats.items(), key=lambda kv: (-kv[1]["count"], kv[0])
    )[:20]
    for (dep_id, arr_id), rst in top_routes:
        dep_ap = rst["dep_airport"]
        arr_ap = rst["arr_airport"]
        dep_label = airport_label(dep_ap)
        arr_label = airport_label(arr_ap)
        base_label = f"{dep_label} \u2192 {arr_label}"
        ws_routes_top.append([base_label, rst["count"]])

    for col in range(1, 3):
        ws_routes_top.column_dimensions[get_column_letter(col)].width = 60

    max_route_count = max((rst["count"] for rst in route_stats.values()), default=0)

    if ws_routes_top.max_row > 1:
        chart = BarChart()
        chart.title = "Count of Routes Traveled"
        chart.legend = None
        configure_chart_axes(chart, x_title="Route", y_title="Count", max_val=max_route_count)
        data = Reference(
            ws_routes_top,
            min_col=2,
            min_row=2,
            max_row=ws_routes_top.max_row,
        )
        cats = Reference(
            ws_routes_top,
            min_col=1,
            min_row=2,
            max_row=ws_routes_top.max_row,
        )
        chart.add_data(data, titles_from_data=False)
        try:
            for s in chart.series:
                s.title = "flights"
        except Exception:
            pass
        chart.set_categories(cats)
        chart.height = 18
        chart.width = 44
        add_value_labels(chart)
        apply_chart_color(chart)
        add_chart_safe(ws_routes_top, chart, "D2")
        ws_routes_top["D22"] = "Routes by leg count"

    # ---------- DayOfWeek sheet ----------
    ws_dow = wb.create_sheet("DayOfWeek")
    ws_dow.append(["idx", "day_name", "flights"])
    for idx in range(7):
        day_name = dt.date(2024, 1, 1 + idx).strftime("%A")
        ws_dow.append([idx, day_name, dow_counts.get(idx, 0)])
    for col in range(1, 4):
        ws_dow.column_dimensions[get_column_letter(col)].width = 16
    chart_dow = BarChart()
    chart_dow.type = "col"
    chart_dow.grouping = "clustered"
    chart_dow.title = "Flights by Day of Week"
    max_dow = max(dow_counts.values()) if dow_counts else 0
    configure_chart_axes(chart_dow, x_title="Day of Week", y_title="Flights", max_val=max_dow)
    chart_dow.legend = None
    data = Reference(ws_dow, min_col=3, min_row=1, max_row=8)
    cats = Reference(ws_dow, min_col=2, min_row=2, max_row=8)
    chart_dow.add_data(data, titles_from_data=True)
    chart_dow.set_categories(cats)
    chart_dow.height = 12
    chart_dow.width = 26
    add_value_labels(chart_dow)
    apply_chart_color(chart_dow)
    add_chart_safe(ws_dow, chart_dow, "E2")
    ws_dow["E16"] = "Flights by day of week"

    # ---------- DayOfMonth sheet ----------
    ws_dom = wb.create_sheet("DayOfMonth")
    ws_dom.append(["day_of_month", "flights"])
    for day in range(1, 32):
        ws_dom.append([day, dom_counts.get(day, 0)])
    for col in range(1, 3):
        ws_dom.column_dimensions[get_column_letter(col)].width = 16
    chart_dom = BarChart()
    chart_dom.type = "col"
    chart_dom.grouping = "clustered"
    chart_dom.title = "Flights by Day of Month"
    max_dom = max(dom_counts.values()) if dom_counts else 0
    configure_chart_axes(chart_dom, x_title="Day of Month", y_title="Flights", max_val=max_dom)
    chart_dom.legend = None
    data = Reference(ws_dom, min_col=2, min_row=1, max_row=32)
    cats = Reference(ws_dom, min_col=1, min_row=2, max_row=32)
    chart_dom.add_data(data, titles_from_data=True)
    chart_dom.set_categories(cats)
    chart_dom.height = 12
    chart_dom.width = 26
    add_value_labels(chart_dom)
    apply_chart_color(chart_dom)
    add_chart_safe(ws_dom, chart_dom, "E2")
    ws_dom["E16"] = "Flights by day of month"

    # ---------- Hours sheet ----------
    ws_hours = wb.create_sheet("Hours")
    ws_hours.append(["hour_utc", "takeoffs", "landings"])
    for h in range(24):
        label = f"{h:02d}00"
        ws_hours.append([label, hour_takeoffs.get(h, 0), hour_landings.get(h, 0)])
    for col in range(1, 4):
        ws_hours.column_dimensions[get_column_letter(col)].width = 16
    chart_hours = BarChart()
    chart_hours.type = "col"
    chart_hours.grouping = "clustered"
    chart_hours.title = "Takeoffs/Landings by Hour of Day"
    max_hour_val = max([0] + list(hour_takeoffs.values()) + list(hour_landings.values()))
    configure_chart_axes(chart_hours, x_title="Hour (UTC)", y_title="Count", max_val=max_hour_val)
    # Legend will be added after data; openpyxl expects a Legend object, not bool
    data = Reference(ws_hours, min_col=2, min_row=1, max_col=3, max_row=25)
    cats = Reference(ws_hours, min_col=1, min_row=2, max_row=25)
    chart_hours.add_data(data, titles_from_data=True)
    chart_hours.set_categories(cats)
    chart_hours.height = 14
    chart_hours.width = 38
    # Enable legend for takeoffs/landings colors
    from openpyxl.chart.legend import Legend
    chart_hours.legend = Legend()
    try:
        # Ensure series names are explicit for legend clarity
        if chart_hours.series and len(chart_hours.series) >= 2:
            chart_hours.series[0].title = "Takeoffs"
            chart_hours.series[1].title = "Landings"
        elif chart_hours.series:
            chart_hours.series[0].title = "Takeoffs"
    except Exception:
        pass
    add_value_labels(chart_hours)
    apply_chart_color(chart_hours, colors=["1f77b4", "ff7f0e"])
    add_chart_safe(ws_hours, chart_hours, "E2")
    ws_hours["E20"] = "Takeoffs / landings by hour (UTC)"

    # ---------- Countries sheet ----------
    ws_countries = wb.create_sheet("Countries")
    ws_countries.append(["country", "visits"])
    country_counts: Dict[str, int] = collections.defaultdict(int)
    for st in airport_stats.values():
        ap = st["airport"]
        name = iso_country_name(ap.get("iso_country"))
        if name:
            country_counts[name] += st["total_visits"]
    max_country = max(country_counts.values()) if country_counts else 0
    for country_name, count in sorted(country_counts.items(), key=lambda kv: (-kv[1], kv[0])):
        ws_countries.append([country_name, count])
    for col in range(1, 3):
        ws_countries.column_dimensions[get_column_letter(col)].width = 18
    if ws_countries.max_row > 1:
        chart_countries = BarChart()
        chart_countries.type = "col"
        chart_countries.grouping = "clustered"
        chart_countries.title = "Country Visits"
        chart_countries.legend = None
        configure_chart_axes(chart_countries, x_title="Country", y_title="Visits", max_val=max_country)
        data = Reference(ws_countries, min_col=2, min_row=1, max_row=ws_countries.max_row)
        cats = Reference(ws_countries, min_col=1, min_row=2, max_row=ws_countries.max_row)
        chart_countries.add_data(data, titles_from_data=True)
        chart_countries.set_categories(cats)
        chart_countries.height = 12
        chart_countries.width = 26
        add_value_labels(chart_countries)
        apply_chart_color(chart_countries)
        add_chart_safe(ws_countries, chart_countries, "D2")
        ws_countries["D18"] = "Visits by country"

    # ---------- Callsigns sheet ----------
    ws_calls = wb.create_sheet("Callsigns")
    ws_calls.append(["callsign", "legs"])
    for cs, count in sorted(
        callsign_counts.items(), key=lambda kv: (-kv[1], kv[0])
    ):
        ws_calls.append([cs, count])

    for col in range(1, 3):
        ws_calls.column_dimensions[get_column_letter(col)].width = 24

    max_callsigns = max(callsign_counts.values()) if callsign_counts else 0

    if ws_calls.max_row > 1:
        chart = BarChart()
        chart.title = "Count of Callsigns"
        chart.legend = None
        configure_chart_axes(chart, x_title="Callsign", y_title="Count", max_val=max_callsigns)
        data = Reference(
            ws_calls, min_col=2, min_row=1, max_row=ws_calls.max_row
        )
        cats = Reference(
            ws_calls, min_col=1, min_row=2, max_row=ws_calls.max_row
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 14
        chart.width = 28
        add_value_labels(chart)
        apply_chart_color(chart)
        add_chart_safe(ws_calls, chart, "E2")
        ws_calls["E20"] = "Legs by callsign"

    # Photos sheet disabled to avoid drawing corruption

    # ---------- Save workbook ----------
    wb.save(out_path)


# -----------------------------
# Meta enrichment – callsign normalization
# -----------------------------


def enrich_meta_from_hits(meta: "AircraftMeta", segments: List[List[Dict[str, Any]]]):
    """
    Look into ac_data inside hits to extract:
      - registration
      - type
      - owner
      - flags (military etc.)
      - callsigns
    """
    callsigns: Set[str] = set(meta.callsigns or [])

    for seg in segments:
        for hit in seg:
            ac = hit.get("ac_data")
            if not isinstance(ac, dict):
                continue

            lower_map = {k.lower(): k for k in ac.keys()}

            def get_ci(*keys):
                for k in keys:
                    real = lower_map.get(k.lower())
                    if real is not None:
                        v = ac.get(real)
                        if isinstance(v, str) and v.strip():
                            return v.strip()
                return None

            if not meta.registration:
                reg = get_ci("r", "reg", "registration", "tail", "tailnum", "tail_num")
                if reg:
                    meta.registration = reg

            if not meta.type:
                t = get_ci("t", "type", "icaoType", "icao_type")
                if t:
                    meta.type = t

            if not meta.description:
                desc = get_ci("desc", "description")
                if desc:
                    meta.description = desc

            if not meta.owner:
                owner = get_ci("owner", "Owner", "op", "operator", "OWNOP")
                if owner:
                    meta.owner = owner

            if not meta.flags:
                dbf_key = lower_map.get("dbflags")
                dbf = ac.get(dbf_key) if dbf_key else None
                f = flags_from_dbflags(dbf) if dbf is not None else None
                if not f:
                    mil = get_ci("MIL", "mil", "military")
                    if mil:
                        f = f"Military={mil}"
                if f:
                    meta.flags = f

            # Callsigns
            for k in ("call", "callsign", "cs", "flight"):
                real = lower_map.get(k.lower())
                if real is None:
                    continue
                v = ac.get(real)
                if isinstance(v, str):
                    v_norm = normalize_callsign(v)
                    if v_norm:
                        callsigns.add(v_norm)

    if callsigns:
        meta.callsigns = sorted(callsigns)


# -----------------------------
# Worker thread
# -----------------------------


class Worker(QtCore.QThread):
    progress = QtCore.pyqtSignal(str)
    card_update = QtCore.pyqtSignal(object)  # AircraftMeta
    finished_ok = QtCore.pyqtSignal()
    finished_err = QtCore.pyqtSignal(str)
    status_update = QtCore.pyqtSignal(str)
    progress_percent = QtCore.pyqtSignal(int)

    def __init__(
        self,
        icao_hex: str,
        start_date: dt.date,
        end_date: dt.date,
        do_kml_points: bool,
        do_kml_routes: bool,
        do_kml_routes3d: bool,
        do_csv: bool,
        do_json: bool,
        do_summary: bool,
        out_dir: str,
    ):
        super().__init__()
        self.icao_hex = icao_hex.lower()
        self.start_date = start_date
        self.end_date = end_date
        self.do_kml_points = do_kml_points
        self.do_kml_routes = do_kml_routes
        self.do_kml_routes3d = do_kml_routes3d
        self.do_csv = do_csv
        self.do_json = do_json
        self.do_summary = do_summary
        self.out_dir = out_dir
        self._stop = False
        self._total_steps = 1
        self._completed_steps = 0
        self.errors: List[str] = []
        self._error_emitted = False

    def _init_progress(self):
        # Total steps: one per day plus one per enabled export
        day_count = (self.end_date - self.start_date).days + 1
        export_steps = sum(
            [
                1 if self.do_kml_points else 0,
                1 if self.do_kml_routes else 0,
                1 if self.do_kml_routes3d else 0,
                1 if self.do_csv else 0,
                1 if self.do_json else 0,
                1 if self.do_summary else 0,
            ]
        )
        total = max(1, day_count + export_steps)
        self._total_steps = total
        self._completed_steps = 0
        self.progress_percent.emit(0)

    def _bump_progress(self):
        self._completed_steps += 1
        pct = int((self._completed_steps / max(1, self._total_steps)) * 100)
        pct = max(0, min(100, pct))
        self.progress_percent.emit(pct)

    def stop(self):
        self._stop = True

    def _emit_error_and_stop(self, msg: str):
        if self._error_emitted:
            return
        self._error_emitted = True
        self._stop = True
        try:
            self.errors.append(msg)
        except Exception:
            pass
        try:
            self.status_update.emit("Stopped")
        except Exception:
            pass
        try:
            self.finished_err.emit(msg)
        except Exception:
            pass

    def _should_abort(self) -> bool:
        return self._stop or self._error_emitted

    def log(self, msg: str):
        self.progress.emit(msg)
        try:
            if any(tag in msg.lower() for tag in ("[error", "[fatal")):
                self.errors.append(msg)
                self._emit_error_and_stop(msg)
        except Exception:
            pass

    def run(self):
        try:
            # 1) Start with just hex
            self.status_update.emit("Querying ...")
            self._init_progress()
            meta_obj = AircraftMeta(hex=self.icao_hex)
            session = requests.Session()

            if self._should_abort():
                return

            # 2) External metadata (best effort, non-fatal)
            try:
                os_meta = fetch_opensky_metadata(self.icao_hex)
                for field in ("registration", "manufacturer", "model", "owner", "type"):
                    v = getattr(os_meta, field, None)
                    if v and not getattr(meta_obj, field):
                        setattr(meta_obj, field, v)
            except Exception as e:
                self.log(f"[meta] OpenSky error: {e}")

            try:
                photo_url, reg2 = fetch_planespotters_photo_and_reg(self.icao_hex)
                if photo_url:
                    meta_obj.photo_url = photo_url
                if reg2 and not meta_obj.registration:
                    meta_obj.registration = reg2
            except Exception as e:
                self.log(f"[meta] Planespotters error: {e}")

            try:
                acdb_root = os.path.join(RESOURCE_ROOT, "acdb_cache")
                db = load_adsbx_acdb(acdb_root, self.log)
                rec = find_acdb_record(db, self.icao_hex)
                if rec:
                    self.log("[acdb] Found record in ADSBx DB")
                    merge_adsbx_record_into_meta(rec, meta_obj)
                else:
                    self.log("[acdb] No record found in ADSBx DB")
            except Exception as e:
                self.log(f"[acdb] Error loading/merging DB: {e}")

            # Normalize type code & friendly name
            apply_type_mapping(meta_obj)

            # Emit initial card
            self.card_update.emit(meta_obj)

            all_segments: List[List[Dict[str, Any]]] = []
            raw_blobs: List[Dict[str, Any]] = []

            # 3) Fetch all days
            for day in daterange(self.start_date, self.end_date):
                if self._should_abort():
                    self.log("[stop] Stopping as requested.")
                    self._emit_error_and_stop("Stopped")
                    return

                self.log(f"[day] {day}")
                cache_root = os.path.join(self.out_dir, "cache")
                path = fetch_trace_for_day(
                    self.icao_hex, day, session, self.log, cache_root=cache_root
                )
                if not path:
                    continue

                try:
                    with open(path, "rb") as f:
                        raw = f.read()
                    try:
                        blob = json.loads(raw)
                    except json.JSONDecodeError:
                        blob = json.loads(
                            gzip.GzipFile(fileobj=io.BytesIO(raw)).read()
                        )
                except Exception as e:
                    self.log(f"[error] parse {day}: {e}")
                    continue

                raw_blobs.append(blob)

                merge_trace_blob_into_meta(blob, meta_obj)
                apply_type_mapping(meta_obj)
                self.card_update.emit(meta_obj)

                segments = extract_hits(blob)
                if segments:
                    all_segments.extend(segments)
                    total_pts = sum(len(seg) for seg in segments)
                    self.log(
                        f"[points] {day}: {total_pts} points in {len(segments)} segment(s)"
                    )
                else:
                    self.log(f"[points] {day}: no valid points found")
                self._bump_progress()
                if self._should_abort():
                    return

            if not all_segments:
                self.log("No points parsed; nothing to write.")
                self.status_update.emit("Complete (no data)")
                self.finished_err.emit("No data")
                return

            # 4) Enrich meta from hits (callsigns, reg, owner etc.)
            enrich_meta_from_hits(meta_obj, all_segments)
            apply_type_mapping(meta_obj)
            self.card_update.emit(meta_obj)

            # 4b) FAA registry (US N-numbers only)
            try:
                if meta_obj.faa_data:
                    self.log("[faa] FAA data already present; skipping lookup.")
                else:
                    reg = meta_obj.registration or ""
                    if not reg.upper().startswith("N"):
                        self.log("[faa] Skipping FAA lookup: no N-number on record.")
                    else:
                        self.log(f"[faa] Looking up FAA registry for {reg} ...")
                        faa_root = os.path.join(RESOURCE_ROOT, "faa_registry")
                        faa_db = load_faa_registry(faa_root, self.log)
                        if faa_db:
                            faa_data = lookup_faa_registry(reg, faa_db)
                            if faa_data:
                                meta_obj.faa_data = faa_data
                                self.log("[faa] FAA registry data loaded")
                            else:
                                self.log(f"[faa] No FAA registry record found for {reg}")
                        else:
                            self.log("[faa] FAA database unavailable")
            except Exception as e:
                self.log(f"[faa] Error: {e}")

            # Basic meta dict for exports
            meta = {
                "hex": meta_obj.hex,
                "registration": meta_obj.registration,
                "type": meta_obj.type,
                "type_name": meta_obj.type_name,
                "owner": meta_obj.owner,
                "manufacturer": meta_obj.manufacturer,
                "model": meta_obj.model,
                "country": meta_obj.country,
                "flags": meta_obj.flags,
                "callsigns": meta_obj.callsigns,
                "description": meta_obj.description,
                "faa": meta_obj.faa_data,
            }

            start_str = self.start_date.strftime("%Y%m%d")
            end_str = self.end_date.strftime("%Y%m%d")
            base_root = os.path.join(
                self.out_dir, f"{self.icao_hex.upper()}_{start_str}_{end_str}"
            )

            try:
                os.makedirs(self.out_dir, exist_ok=True)
            except Exception:
                pass

            total_points = sum(len(seg) for seg in all_segments)
            self.log(f"[stats] Total points across all days: {total_points}")
            self.status_update.emit("Generating KML ...")
            if self._should_abort():
                return

            # Precompute legs for KML if needed
            legs: List[Dict[str, Any]] = []
            if self.do_kml_points or self.do_kml_routes or self.do_kml_routes3d or self.do_summary:
                self.log("[kml] Extracting legs…")
                legs = extract_legs_for_kml(all_segments)
                self.log(f"[kml] Found {len(legs)} leg(s)")

            airports_normalized: List[Dict[str, Any]] = []
            legs_for_kml = legs
            if (
                legs
                and (
                    self.do_kml_points
                    or self.do_kml_routes
                    or self.do_kml_routes3d
                    or self.do_summary
                )
            ):
                try:
                    ap_root = os.path.join(RESOURCE_ROOT, "airport_db")
                    airport_db_for_kml = load_airport_db_for_summary(ap_root, self.log)
                    airports_normalized = normalize_airports_for_matching(airport_db_for_kml)
                    if airports_normalized:
                        legs_for_kml, _, _, _ = augment_legs_with_airports(legs, airports_normalized)
                except Exception as e:
                    self.log(f"[airports] Could not load/augment airports for KML: {e}")

            # KML: points
            if self.do_kml_points:
                if self._should_abort():
                    self.log("[stop] Stopping before KML points as requested.")
                    self._emit_error_and_stop("Stopped")
                    return
                kml_points_path = base_root + "_points.kml"
                self.log(f"[kml] Building points KML at {kml_points_path}…")
                try:
                    build_kml_points(legs_for_kml, self.icao_hex, meta, kml_points_path)
                    self.log(f"[kml] Wrote {kml_points_path}")
                except Exception as e:
                    self.log(f"[error] KML points: {e}")
                self._bump_progress()

            # KML: routes 2D
            if self.do_kml_routes:
                if self._should_abort():
                    self.log("[stop] Stopping before KML routes as requested.")
                    self._emit_error_and_stop("Stopped")
                    return
                kml_routes_path = base_root + "_routes.kml"
                self.log(f"[kml] Building routes KML at {kml_routes_path}…")
                try:
                    build_kml_routes_2d(legs_for_kml, self.icao_hex, meta, kml_routes_path)
                    self.log(f"[kml] Wrote {kml_routes_path}")
                except Exception as e:
                    self.log(f"[error] KML routes: {e}")
                self._bump_progress()

            # KML: routes 3D
            if self.do_kml_routes3d:
                if self._should_abort():
                    self.log("[stop] Stopping before KML routes 3D as requested.")
                    self._emit_error_and_stop("Stopped")
                    return
                kml_routes3d_path = base_root + "_routes3d.kml"
                self.log(f"[kml] Building 3D routes KML at {kml_routes3d_path}…")
                try:
                    build_kml_routes_3d(legs_for_kml, self.icao_hex, meta, kml_routes3d_path)
                    self.log(f"[kml] Wrote {kml_routes3d_path}")
                except Exception as e:
                    self.log(f"[error] KML routes 3D: {e}")
                self._bump_progress()

            # CSV
            if self.do_csv:
                if self._should_abort():
                    self.log("[stop] Stopping before CSV as requested.")
                    self._emit_error_and_stop("Stopped")
                    return
                self.status_update.emit("Generating CSV ...")
                csv_path = base_root + ".csv"
                self.log(f"[csv] Building CSV at {csv_path}…")
                try:
                    build_csv(all_segments, self.icao_hex, meta, csv_path)
                    self.log(f"[csv] Wrote {csv_path}")
                except Exception as e:
                    self.log(f"[error] CSV: {e}")
                self._bump_progress()

            # JSON
            if self.do_json:
                if self._should_abort():
                    self.log("[stop] Stopping before JSON as requested.")
                    self._emit_error_and_stop("Stopped")
                    return
                self.status_update.emit("Generating JSON ...")
                json_path = base_root + ".json"
                self.log(f"[json] Building JSON at {json_path}…")
                try:
                    build_json(raw_blobs, json_path)
                    self.log(f"[json] Wrote {json_path}")
                except Exception as e:
                    self.log(f"[error] JSON: {e}")
                self._bump_progress()

            # Excel summary
            if self.do_summary:
                if self._should_abort():
                    self.log("[stop] Stopping before summary as requested.")
                    self._emit_error_and_stop("Stopped")
                    return
                self.status_update.emit("Generating Summary ...")
                try:
                    ap_root = os.path.join(RESOURCE_ROOT, "airport_db")
                    airport_db = load_airport_db_for_summary(ap_root, self.log)
                    if airport_db is not None:
                        summary_path = base_root + "_summary.xlsx"
                        self.log(f"[summary] Building Excel summary at {summary_path}…")
                        build_summary_excel(
                            all_segments,
                            self.icao_hex,
                            meta_obj,
                            self.start_date,
                            self.end_date,
                            summary_path,
                            airport_db,
                        )
                        self.log(f"[summary] Wrote {summary_path}")
                    else:
                        self._emit_error_and_stop("Airport database not available.")
                        return
                except Exception as e:
                    self.log(f"[error] Summary: {e}")
                    self._emit_error_and_stop(f"Summary: {e}")
                    return
                self._bump_progress()

            # Clear per-hex cache after a successful run
            try:
                cache_root = os.path.join(self.out_dir, "cache")
                cache_hex_dir = os.path.join(cache_root, self.icao_hex.lower())
                if os.path.isdir(cache_hex_dir):
                    shutil.rmtree(cache_hex_dir, ignore_errors=False)
                    self.log(f"[cache] Cleared {cache_hex_dir}")
                # Remove empty cache root
                if os.path.isdir(cache_root) and not os.listdir(cache_root):
                    shutil.rmtree(cache_root, ignore_errors=False)
            except Exception as e:
                self.log(f"[cache] Unable to clear cache: {e}")

            # Emit any collected errors at the end
            if self.errors:
                self.log("[errors] Summary of errors during run:")
                for err in self.errors:
                    self.log(err)

            self.finished_ok.emit()
            self.status_update.emit("Complete")
            self.progress_percent.emit(100)

        except Exception as e:
            self.log(f"[fatal] {e}")
            self._emit_error_and_stop(str(e))
            return


# -----------------------------
# UI helpers and MainWindow
# -----------------------------


class ToggleCheckBox(QtWidgets.QCheckBox):
    """
    Simple toggle-style checkbox that turns green when checked.
    """

    def __init__(self, text: str = "", parent: Optional[QtWidgets.QWidget] = None):
        super().__init__(text, parent)
        self.stateChanged.connect(self._update_style)
        self._update_style()

    def _update_style(self):
        if self.isChecked():
            self.setStyleSheet(
                """
                QCheckBox {
                    background-color: #2e7d32;
                    color: white;
                    border-radius: 4px;
                    padding: 4px 8px;
                }
                """
            )
        else:
            self.setStyleSheet(
                """
                QCheckBox {
                    background-color: #424242;
                    color: #e0e0e0;
                    border-radius: 4px;
                    padding: 4px 8px;
                }
                """
            )


class ImageLabel(QtWidgets.QLabel):
    """
    QLabel that can load and display an image from a URL.
    """

    def __init__(self, parent: Optional[QtWidgets.QWidget] = None):
        super().__init__(parent)
        self.setMinimumHeight(200)
        self.setAlignment(QtCore.Qt.AlignCenter)
        self.setScaledContents(False)
        self._orig_pixmap: Optional[QtGui.QPixmap] = None

    def set_image_from_url(self, url: Optional[str]):
        if isinstance(url, dict):
            url = url.get("src")
        if not url:
            self.clear()
            self.setText("No image")
            return
        try:
            resp = requests.get(
                url,
                timeout=10,
                headers={"User-Agent": "adsbx-history-downloader-gui/1.0"},
            )
            if resp.status_code != 200:
                self.clear()
                self.setText("No image")
                return
            img_data = resp.content
            pixmap = QtGui.QPixmap()
            if not pixmap.loadFromData(img_data):
                self.clear()
                self.setText("No image")
                return
            self._orig_pixmap = pixmap
            # Fit into label while preserving aspect ratio
            scaled = pixmap.scaled(
                self.width() or 400,
                self.height() or 250,
                QtCore.Qt.KeepAspectRatio,
                QtCore.Qt.SmoothTransformation,
            )
            self.setPixmap(scaled)
        except Exception:
            self.clear()
            self.setText("No image")

    def resizeEvent(self, event: QtGui.QResizeEvent) -> None:
        # Keep image scaled when the label is resized
        src = self._orig_pixmap or self.pixmap()
        if src is not None and not src.isNull():
            scaled = src.scaled(
                self.width(),
                self.height(),
                QtCore.Qt.KeepAspectRatio,
                QtCore.Qt.SmoothTransformation,
            )
            self.setPixmap(scaled)
        super().resizeEvent(event)


class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("SkyProfile")
        try:
            ico = QtGui.QIcon(resource_path("adsbtrack_icon_orange.ico"))
            if not ico.isNull():
                self.setWindowIcon(ico)
                app = QtWidgets.QApplication.instance()
                if app:
                    app.setWindowIcon(ico)
        except Exception:
            pass
        self.resize(int(1100 * 1.2), 700)

        self.worker: Optional[Worker] = None

        central = QtWidgets.QWidget()
        self.setCentralWidget(central)
        vbox = QtWidgets.QVBoxLayout(central)

        splitter = QtWidgets.QSplitter(QtCore.Qt.Horizontal)

        # Left pane: inputs
        left = QtWidgets.QWidget()
        form = QtWidgets.QFormLayout(left)
        form.setFieldGrowthPolicy(QtWidgets.QFormLayout.AllNonFixedFieldsGrow)

        # Keep inputs at a single-line height for a tighter, consistent form row
        def set_single_line_height(widget: QtWidgets.QWidget, padding: int = 4):
            fm = QtGui.QFontMetrics(widget.font())
            h = fm.height() + padding
            widget.setMinimumHeight(h)
            widget.setMaximumHeight(h)
            sp = widget.sizePolicy()
            sp.setVerticalPolicy(QtWidgets.QSizePolicy.Fixed)
            widget.setSizePolicy(sp)

        def set_field_width(widget: QtWidgets.QWidget, width: int = 400):
            widget.setMinimumWidth(width)
            widget.setMaximumWidth(width)

        self.start_date = QtWidgets.QDateEdit()
        self.start_date.setCalendarPopup(True)
        self.start_date.setDate(QtCore.QDate.currentDate().addDays(-1))
        set_single_line_height(self.start_date)
        set_field_width(self.start_date)

        self.end_date = QtWidgets.QDateEdit()
        self.end_date.setCalendarPopup(True)
        self.end_date.setDate(QtCore.QDate.currentDate())
        set_single_line_height(self.end_date)
        set_field_width(self.end_date)

        self.hex_edit = QtWidgets.QLineEdit()
        self.hex_edit.setPlaceholderText("ICAO HEX (e.g., A1B2C3)")
        self.hex_edit.textChanged.connect(self._upper_hex)
        set_single_line_height(self.hex_edit)
        set_field_width(self.hex_edit)

        # Optional tail / registration lookup
        self.tail_edit = QtWidgets.QLineEdit()
        self.tail_edit.setPlaceholderText("Tail / Registration (optional)")
        set_single_line_height(self.tail_edit)
        set_field_width(self.tail_edit)

        # Export toggles
        self.kml_points_chk = ToggleCheckBox("Points KML")
        self.kml_routes_chk = ToggleCheckBox("Routes KML")
        self.kml_routes3d_chk = ToggleCheckBox("3D Routes KML")
        self.csv_chk = ToggleCheckBox("Export CSV")
        self.json_chk = ToggleCheckBox("Export JSON")
        self.summary_chk = ToggleCheckBox("Export Excel Summary")

        # Defaults: CSV/JSON/Summary on, KML points+routes on, 3D routes off
        self.kml_points_chk.setChecked(True)
        self.kml_routes_chk.setChecked(True)
        self.kml_routes3d_chk.setChecked(False)
        self.csv_chk.setChecked(True)
        self.json_chk.setChecked(True)
        self.summary_chk.setChecked(True)

        # Output folder row: path + Browse + Open
        out_box = QtWidgets.QHBoxLayout()
        self.out_edit = QtWidgets.QLineEdit()
        self.out_btn = QtWidgets.QPushButton("Browse...")
        self.out_btn.clicked.connect(self.choose_folder)
        self.out_open_btn = QtWidgets.QPushButton("Open")
        self.out_open_btn.clicked.connect(self.open_folder_in_explorer)
        set_field_width(self.out_edit, width=400)
        set_single_line_height(self.out_edit)
        set_single_line_height(self.out_btn)
        set_single_line_height(self.out_open_btn)
        self.out_btn.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        self.out_open_btn.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        out_box.setSpacing(int(out_box.spacing() * 0.75))
        out_box.setContentsMargins(0, 0, 0, 0)
        out_box.addWidget(self.out_edit, 1)
        out_box.addWidget(self.out_btn)
        out_box.addWidget(self.out_open_btn)
        out_wrap = QtWidgets.QWidget()
        out_wrap.setLayout(out_box)

        self.run_btn = QtWidgets.QPushButton("Run Query")
        self.stop_btn = QtWidgets.QPushButton("Stop Query")
        self.run_btn.clicked.connect(self.run_query)
        self.stop_btn.clicked.connect(self.stop_query)
        btn_style = """
            QPushButton {
                background-color: #555555;
                color: white;
                border: 1px solid #555555;
                border-radius: 4px;
                padding: 4px 10px;
            }
            QPushButton:pressed {
                background-color: #4a4a4a;
            }
            QPushButton:focus {
                outline: 2px solid #777777;
            }
            QPushButton:disabled {
                background-color: #4f4f4f;
                color: #bdbdbd;
                border-color: #4f4f4f;
            }
        """
        self.run_btn.setStyleSheet(btn_style)
        self.stop_btn.setStyleSheet(btn_style)
        for btn in (self.run_btn, self.stop_btn):
            size = btn.sizeHint()
            btn.setFixedSize(int(size.width() * 0.75), int(size.height() * 0.75))
        btns = QtWidgets.QHBoxLayout()
        btns.setSpacing(int(btns.spacing() * 0.75))
        btns.setContentsMargins(0, 0, 0, 0)
        btns.addWidget(self.run_btn)
        btns.addWidget(self.stop_btn)
        btn_wrap = QtWidgets.QWidget()
        btn_wrap.setLayout(btns)

        self.current_status = "Idle"
        self.progress_bar = QtWidgets.QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("Idle")
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setStyleSheet(
            """
            QProgressBar {
                border: 1px solid #2e7d32;
                border-radius: 4px;
                background: #2a2a2a;
                color: white;
                text-align: center;
            }
            QProgressBar::chunk {
                background-color: #2e7d32;
            }
            """
        )

        form.addRow("Start date:", self.start_date)
        form.addRow("Stop date:", self.end_date)
        form.addRow("ICAO HEX:", self.hex_edit)
        form.addRow("Tail / Reg:", self.tail_edit)

        # KML outputs grouped and slightly indented
        kml_vbox = QtWidgets.QVBoxLayout()
        kml_vbox.setContentsMargins(0, 0, 0, 0)
        kml_vbox.addWidget(self.kml_points_chk)
        kml_vbox.addWidget(self.kml_routes_chk)
        kml_vbox.addWidget(self.kml_routes3d_chk)
        kml_wrap = QtWidgets.QWidget()
        kml_wrap.setLayout(kml_vbox)
        form.addRow(kml_wrap)

        form.addRow(self.csv_chk)
        form.addRow(self.json_chk)
        form.addRow(self.summary_chk)
        form.addRow("Output folder:", out_wrap)
        form.addRow(btn_wrap)
        form.addRow("", self.progress_bar)

        splitter.addWidget(left)

        # Right pane: baseball card + log
        right = QtWidgets.QWidget()
        right_layout = QtWidgets.QVBoxLayout(right)

        # Baseball card group
        card_group = QtWidgets.QGroupBox("Aircraft")
        card_layout = QtWidgets.QGridLayout(card_group)

        self.label_hex_val = QtWidgets.QLabel("—")
        self.label_reg_val = QtWidgets.QLabel("—")
        self.label_type_val = QtWidgets.QLabel("—")
        self.label_typename_val = QtWidgets.QLabel("—")
        self.label_owner_val = QtWidgets.QLabel("—")
        self.label_mfr_val = QtWidgets.QLabel("—")
        self.label_model_val = QtWidgets.QLabel("—")
        self.label_flags_val = QtWidgets.QLabel("—")

        row = 0
        card_layout.addWidget(QtWidgets.QLabel("ICAO Hex:"), row, 0)
        card_layout.addWidget(self.label_hex_val, row, 1)
        row += 1
        card_layout.addWidget(QtWidgets.QLabel("Registration:"), row, 0)
        card_layout.addWidget(self.label_reg_val, row, 1)
        row += 1
        card_layout.addWidget(QtWidgets.QLabel("Type:"), row, 0)
        card_layout.addWidget(self.label_type_val, row, 1)
        row += 1
        card_layout.addWidget(QtWidgets.QLabel("Type name:"), row, 0)
        card_layout.addWidget(self.label_typename_val, row, 1)
        row += 1
        card_layout.addWidget(QtWidgets.QLabel("Owner:"), row, 0)
        card_layout.addWidget(self.label_owner_val, row, 1)
        row += 1
        card_layout.addWidget(QtWidgets.QLabel("Manufacturer:"), row, 0)
        card_layout.addWidget(self.label_mfr_val, row, 1)
        row += 1
        card_layout.addWidget(QtWidgets.QLabel("Model:"), row, 0)
        card_layout.addWidget(self.label_model_val, row, 1)
        row += 1
        card_layout.addWidget(QtWidgets.QLabel("Flags:"), row, 0)
        card_layout.addWidget(self.label_flags_val, row, 1)

        right_layout.addWidget(card_group)

        # Photo
        self.photo_label = ImageLabel()
        right_layout.addWidget(self.photo_label)

        # Log box
        self.log = QtWidgets.QTextEdit()
        self.log.setReadOnly(True)
        self.log.setLineWrapMode(QtWidgets.QTextEdit.NoWrap)
        font = QtGui.QFontDatabase.systemFont(QtGui.QFontDatabase.FixedFont)
        self.log.setFont(font)
        right_layout.addWidget(self.log)

        splitter.addWidget(right)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 2)

        vbox.addWidget(splitter)

        self._apply_dark_theme()

    # ---- UI helpers ----

    def _apply_dark_theme(self):
        app = QtWidgets.QApplication.instance()
        if not app:
            return
        palette = QtGui.QPalette()
        palette.setColor(QtGui.QPalette.Window, QtGui.QColor(53, 53, 53))
        palette.setColor(QtGui.QPalette.WindowText, QtCore.Qt.white)
        palette.setColor(QtGui.QPalette.Base, QtGui.QColor(25, 25, 25))
        palette.setColor(QtGui.QPalette.AlternateBase, QtGui.QColor(53, 53, 53))
        palette.setColor(QtGui.QPalette.ToolTipBase, QtCore.Qt.white)
        palette.setColor(QtGui.QPalette.ToolTipText, QtCore.Qt.white)
        palette.setColor(QtGui.QPalette.Text, QtCore.Qt.white)
        palette.setColor(QtGui.QPalette.Button, QtGui.QColor(53, 53, 53))
        palette.setColor(QtGui.QPalette.ButtonText, QtCore.Qt.white)
        palette.setColor(QtGui.QPalette.BrightText, QtCore.Qt.red)
        palette.setColor(QtGui.QPalette.Highlight, QtGui.QColor(142, 45, 197).lighter())
        palette.setColor(QtGui.QPalette.HighlightedText, QtCore.Qt.black)
        app.setPalette(palette)

    def _upper_hex(self, txt: str):
        self.hex_edit.blockSignals(True)
        self.hex_edit.setText(txt.upper())
        self.hex_edit.blockSignals(False)

    def choose_folder(self):
        path = QtWidgets.QFileDialog.getExistingDirectory(self, "Choose output folder")
        if path:
            self.out_edit.setText(path)

    def open_folder_in_explorer(self):
        path = self.out_edit.text().strip()
        if not path:
            return
        if sys.platform.startswith("win"):
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])

    def append_log(self, text: str):
        ts = dt.datetime.now(dt.timezone.utc).strftime("%H:%M:%S")
        line = f"[{ts}] {text}"
        is_error = any(tag in text.lower() for tag in ("[error", "[fatal"))
        line_safe = html.escape(line)
        if is_error:
            self.log.append(f'<span style="color:#e53935">{line_safe}</span>')
        else:
            self.log.append(line_safe)

    def on_status_update(self, text: str):
        self.current_status = text

    def on_progress_percent(self, pct: int):
        try:
            self.progress_bar.setValue(max(0, min(100, int(pct))))
            pct_clean = int(max(0, min(100, pct)))
            self.progress_bar.setFormat(f"{self.current_status} ({pct_clean}%)")
        except Exception:
            pass

    def on_query_finished_ok(self):
        self.append_log("[done] All days processed (consolidated).")
        self.current_status = "Complete"
        self.progress_bar.setValue(100)
        self.progress_bar.setFormat("Query Complete")
        self.run_btn.setText("Run Query")
        self.run_btn.setEnabled(True)
        out_dir = self.out_edit.text().strip()

        msg = QtWidgets.QMessageBox(self)
        msg.setWindowTitle("Query complete")
        body = "Query complete."
        if out_dir:
            body += f"\nOutput folder:\n{out_dir}"
        msg.setText(body)

        open_btn = msg.addButton("Open Folder", QtWidgets.QMessageBox.ActionRole)
        msg.addButton(QtWidgets.QMessageBox.Ok)
        msg.exec_()

        if msg.clickedButton() == open_btn and out_dir and os.path.isdir(out_dir):
            self.open_folder_in_explorer()

    def on_card_update(self, meta_obj: AircraftMeta):
        self.label_hex_val.setText(meta_obj.hex.upper())
        self.label_reg_val.setText(meta_obj.registration or "—")
        self.label_type_val.setText(meta_obj.type or "—")
        self.label_typename_val.setText(meta_obj.type_name or "—")
        self.label_owner_val.setText(meta_obj.owner or "—")
        self.label_mfr_val.setText(meta_obj.manufacturer or "—")
        self.label_model_val.setText(meta_obj.model or "—")
        self.label_flags_val.setText(meta_obj.flags or "–")
        self.photo_label.set_image_from_url(meta_obj.photo_url)

    def run_query(self):
        if self.worker is not None and self.worker.isRunning():
            self.append_log("[busy] A query is already running.")
            return

        hex_code = self.hex_edit.text().strip().upper()
        tail = self.tail_edit.text().strip().upper()

        # If ICAO hex is missing or invalid but a tail is provided, try lookup
        def is_valid_hex(h: str) -> bool:
            return bool(h) and all(c in "0123456789ABCDEF" for c in h)

        if not is_valid_hex(hex_code):
            if not tail:
                QtWidgets.QMessageBox.warning(
                    self,
                    "Missing identifier",
                    "Please enter a valid ICAO hex or a tail/registration.",
                )
                return

            # Try to resolve hex from ADSBx aircraft DB by registration
            out_dir = self.out_edit.text().strip() or os.getcwd()
            acdb_root = os.path.join(RESOURCE_ROOT, "acdb_cache")
            self.append_log(f"[lookup] Resolving tail {tail} via ADSBx DB…")
            try:
                db = load_adsbx_acdb(acdb_root, self.append_log)
                result = find_acdb_record_by_reg(db, tail)
            except Exception as e:
                QtWidgets.QMessageBox.warning(
                    self,
                    "Tail lookup failed",
                    f"Error loading ADSBx aircraft DB: {e}",
                )
                return

            if not result:
                QtWidgets.QMessageBox.information(
                    self,
                    "Tail not found",
                    f"Could not find ICAO hex for tail {tail} in ADSBx DB.",
                )
                return

            resolved_hex, _rec = result
            hex_code = resolved_hex.upper()
            self.hex_edit.setText(hex_code)
            self.append_log(f"[lookup] Tail {tail} → ICAO hex {hex_code}")

        sdate = self.start_date.date().toPyDate()
        edate = self.end_date.date().toPyDate()
        if edate < sdate:
            sdate, edate = edate, sdate

        out_dir = self.out_edit.text().strip()
        if not out_dir:
            QtWidgets.QMessageBox.warning(
                self, "Output folder", "Please choose an output folder."
            )
            return

        do_kml_points = self.kml_points_chk.isChecked()
        do_kml_routes = self.kml_routes_chk.isChecked()
        do_kml_routes3d = self.kml_routes3d_chk.isChecked()
        do_csv = self.csv_chk.isChecked()
        do_json = self.json_chk.isChecked()
        do_summary = self.summary_chk.isChecked()

        if not any(
            [do_kml_points, do_kml_routes, do_kml_routes3d, do_csv, do_json, do_summary]
        ):
            QtWidgets.QMessageBox.information(
                self, "Nothing to export", "Enable at least one export format."
            )
            return

        os.makedirs(out_dir, exist_ok=True)

        self.log.clear()
        self.append_log(
            f"[run] HEX={hex_code}, {sdate} → {edate}, out={out_dir}"
        )
        self.run_btn.setText("Running")
        self.run_btn.setEnabled(False)
        self.current_status = "Querying ..."
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("Querying ... (0%)")
        self.worker = Worker(
            hex_code,
            sdate,
            edate,
            do_kml_points,
            do_kml_routes,
            do_kml_routes3d,
            do_csv,
            do_json,
            do_summary,
            out_dir,
        )
        self.worker.progress.connect(self.append_log)
        self.worker.card_update.connect(self.on_card_update)
        self.worker.finished_ok.connect(self.on_query_finished_ok)
        self.worker.finished_err.connect(self.on_query_finished_err)
        self.worker.status_update.connect(self.on_status_update)
        self.worker.progress_percent.connect(self.on_progress_percent)
        self.worker.start()

    def stop_query(self):
        if self.worker is not None and self.worker.isRunning():
            self.append_log("[req] Stop requested…")
            self.current_status = "Stopping ..."
            self.progress_bar.setFormat("Stopping ...")
            self.worker.stop()
        else:
            self.append_log("[idle] No running query.")

    def on_query_finished_err(self, message: str):
        self.append_log(f"[stopped] {message}")
        if message.lower() == "stopped":
            self.current_status = "Stopped"
            self.progress_bar.setFormat("Stopped")
        else:
            self.current_status = "Complete"
            self.progress_bar.setFormat("Complete")
        self.progress_bar.setValue(100)
        self.run_btn.setText("Run Query")
        self.run_btn.setEnabled(True)

        # Show failure dialog with error message
        msg_box = QtWidgets.QMessageBox(self)
        msg_box.setWindowTitle("Failed")
        msg_box.setText(f"Failed: {message}")
        msg_box.setIcon(QtWidgets.QMessageBox.Critical)
        msg_box.addButton(QtWidgets.QMessageBox.Ok)
        msg_box.exec_()


def main():
    app = QtWidgets.QApplication(sys.argv)
    app.setStyle("Fusion")
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
